import os

BASE = "humix"
dirs = [
    f"{BASE}/templates",
    f"{BASE}/static/css",
    f"{BASE}/static/js",
    f"{BASE}/static/captured_images",
    f"{BASE}/database",
    f"{BASE}/arduino",
]
for d in dirs:
    os.makedirs(d, exist_ok=True)

files = {}

# ── REQUIREMENTS ──────────────────────────────────────────
files[f"{BASE}/requirements.txt"] = """\
flask
opencv-python
ultralytics
numpy
openpyxl
Pillow
pyserial
"""

# ── APP.PY ────────────────────────────────────────────────
files[f"{BASE}/app.py"] = r'''
from flask import (Flask, render_template, jsonify, request,
                   session, redirect, url_for, Response, send_file)
import cv2
from ultralytics import YOLO
import time, threading, numpy as np, urllib.request
import sqlite3, os, io, uuid, hashlib
from datetime import datetime
from functools import wraps

app = Flask(__name__)
app.secret_key = "humix-secret-2024"

ADMIN_USER     = "admin"
ADMIN_PASS     = "admin123"
DB_PATH        = "database/humix.db"
CAPTURE_INTV   = 2    # seconds between snapshots
IMAGE_INTV     = 180  # 3 minutes image save

# ── GLOBAL STATE ────────────────────────────────────────
model             = None
camera            = None
camera_active     = False
camera_source     = "webcam"
esp32_url         = ""
current_count     = 0
alert_active      = False
flow_direction    = "Analyzing..."
latest_frame      = None
frame_lock        = threading.Lock()
capture_thread    = None
current_event_id  = None
current_threshold = 10
prev_gray         = None
last_image_save   = 0
arduino_serial    = None

session_data = {
    "start_time": None, "total_logs": 0,
    "alert_events": 0,  "max_crowd": 0,
    "session_id": None,
}

# ── DATABASE ────────────────────────────────────────────
def init_db():
    os.makedirs("database", exist_ok=True)
    os.makedirs(os.path.join("static","captured_images"), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_name TEXT, date TEXT, start_time TEXT, end_time TEXT,
        venue TEXT, estimated_crowd INTEGER, organizer_name TEXT,
        email TEXT UNIQUE, password TEXT,
        status TEXT DEFAULT 'registered',
        created_at TEXT);
    CREATE TABLE IF NOT EXISTS crowd_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_id INTEGER, session_id TEXT, timestamp TEXT,
        count INTEGER, threshold INTEGER, alert INTEGER, flow_direction TEXT);
    CREATE TABLE IF NOT EXISTS saved_images (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_id INTEGER, session_id TEXT, timestamp TEXT,
        filename TEXT, count INTEGER);
    CREATE TABLE IF NOT EXISTS mon_sessions (
        id TEXT PRIMARY KEY, event_id INTEGER,
        start_time TEXT, end_time TEXT,
        max_crowd INTEGER, total_alerts INTEGER,
        total_logs INTEGER, threshold INTEGER);
    """)
    conn.commit(); conn.close()

def get_db():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c

def hp(p): return hashlib.sha256(p.encode()).hexdigest()

# ── ARDUINO ─────────────────────────────────────────────
def connect_arduino(port):
    global arduino_serial
    try:
        import serial
        arduino_serial = serial.Serial(port, 9600, timeout=1)
        time.sleep(2)
        send_led("G")
        print(f"[HUMIX] Arduino on {port}")
        return True
    except Exception as e:
        print(f"[HUMIX] Arduino error: {e}")
        arduino_serial = None
        return False

def send_led(cmd):
    try:
        if arduino_serial and arduino_serial.is_open:
            arduino_serial.write(f"{cmd}\n".encode())
    except:
        pass

def update_leds(count, threshold):
    if count >= threshold:
        send_led("R")
    elif count >= threshold * 0.75:
        send_led("Y")
    else:
        send_led("G")

# ── MODEL ───────────────────────────────────────────────
def load_model():
    global model
    if model is None:
        print("[HUMIX] Loading model...")
        try:
            model = YOLO("yolo12n.pt")
            print("[HUMIX] YOLOv12 loaded!")
        except:
            model = YOLO("yolov8n.pt")
            print("[HUMIX] YOLOv8 loaded (fallback)")

# ── PREPROCESS ──────────────────────────────────────────
def preprocess(frame):
    frame = cv2.resize(frame, (640, 480))
    lab = cv2.cvtColor(frame, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    l = cv2.equalizeHist(l)
    frame = cv2.cvtColor(cv2.merge([l,a,b]), cv2.COLOR_LAB2BGR)
    blurred = cv2.GaussianBlur(frame,(0,0),3)
    return cv2.addWeighted(frame,1.5,blurred,-0.5,0)

# ── OPTICAL FLOW ────────────────────────────────────────
def get_flow(frame):
    global prev_gray
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    d = "Still"
    if prev_gray is not None and prev_gray.shape == gray.shape:
        try:
            flow = cv2.calcOpticalFlowFarneback(
                prev_gray,gray,None,0.5,3,15,3,5,1.2,0)
            fx,fy = flow[...,0].mean(), flow[...,1].mean()
            mag = (fx**2+fy**2)**0.5
            if mag > 0.3:
                if abs(fx)>abs(fy): d="Moving RIGHT →" if fx>0 else "Moving LEFT ←"
                else:               d="Moving DOWN ↓"  if fy>0 else "Moving UP ↑"
        except: d="Analyzing..."
    prev_gray = gray.copy()
    return d

# ── MJPEG STREAM (event page) ───────────────────────────
def generate_stream():
    while camera_active:
        if camera_source=="webcam" and camera and camera.isOpened():
            ok, frame = camera.read()
            if ok and frame is not None:
                cv2.putText(frame,f"People: {current_count}",(10,30),
                    cv2.FONT_HERSHEY_SIMPLEX,0.8,(0,220,100),2)
                cv2.putText(frame,f"Threshold: {current_threshold}",(10,58),
                    cv2.FONT_HERSHEY_SIMPLEX,0.6,(255,255,255),1)
                cv2.putText(frame,flow_direction,(10,86),
                    cv2.FONT_HERSHEY_SIMPLEX,0.55,(200,200,0),1)
                if alert_active:
                    cv2.rectangle(frame,(0,0),(frame.shape[1]-1,frame.shape[0]-1),(0,0,255),4)
                    cv2.putText(frame,"ALERT: CROWD LIMIT EXCEEDED",
                        (10,frame.shape[0]-15),cv2.FONT_HERSHEY_SIMPLEX,0.65,(0,0,255),2)
                ret,buf = cv2.imencode(".jpg",frame,[cv2.IMWRITE_JPEG_QUALITY,70])
                if ret:
                    yield (b"--frame\r\nContent-Type: image/jpeg\r\n\r\n"
                           +buf.tobytes()+b"\r\n")
        else:
            with frame_lock:
                f = latest_frame.copy() if latest_frame is not None else None
            if f is not None:
                ret,buf = cv2.imencode(".jpg",f,[cv2.IMWRITE_JPEG_QUALITY,70])
                if ret:
                    yield (b"--frame\r\nContent-Type: image/jpeg\r\n\r\n"
                           +buf.tobytes()+b"\r\n")
        time.sleep(0.033)

# ── CAPTURE WORKER ──────────────────────────────────────
def capture_worker():
    global current_count,alert_active,latest_frame,flow_direction
    global session_data,last_image_save,current_threshold
    last_capture=0; last_image_save=time.time()

    while camera_active:
        now=time.time()
        if last_capture!=0 and (now-last_capture)<CAPTURE_INTV:
            time.sleep(0.05); continue

        raw=None
        try:
            if camera_source=="webcam":
                if camera and camera.isOpened():
                    for _ in range(3): ok,raw=camera.read()
                    if not ok or raw is None: time.sleep(0.5); continue
                    if raw.mean()<10:          time.sleep(0.3); continue
            elif camera_source=="esp32" and esp32_url:
                resp=urllib.request.urlopen(esp32_url,timeout=5)
                arr=np.frombuffer(resp.read(),dtype=np.uint8)
                raw=cv2.imdecode(arr,cv2.IMREAD_COLOR)
                if raw is None: time.sleep(0.5); continue
        except Exception as e:
            print(f"[HUMIX] Grab: {e}"); time.sleep(0.5); continue

        if raw is None: continue

        try: frame=preprocess(raw)
        except: frame=raw.copy()

        try: flow_direction=get_flow(frame)
        except: flow_direction="Analyzing..."

        count=0; display=frame.copy()
        try:
            for r in model(frame,stream=True,conf=0.5,verbose=False):
                for box in r.boxes:
                    if int(box.cls[0])==0:
                        count+=1
                        x1,y1,x2,y2=map(int,box.xyxy[0])
                        cv2.rectangle(display,(x1,y1),(x2,y2),(0,220,100),2)
                        cv2.putText(display,f"{float(box.conf[0]):.2f}",
                            (x1,y1-6),cv2.FONT_HERSHEY_SIMPLEX,0.45,(0,220,100),1)
        except Exception as e: print(f"[HUMIX] Detect: {e}")

        prev=alert_active
        current_count=count
        alert_active=count>current_threshold
        if alert_active and not prev: session_data["alert_events"]+=1
        session_data["max_crowd"]=max(session_data["max_crowd"],count)
        session_data["total_logs"]+=1

        update_leds(count,current_threshold)

        sid=session_data["session_id"]
        if current_event_id and sid:
            try:
                conn=get_db()
                conn.execute("INSERT INTO crowd_logs VALUES (NULL,?,?,?,?,?,?,?)",
                    (current_event_id,sid,datetime.now().strftime("%H:%M:%S"),
                     count,current_threshold,1 if alert_active else 0,flow_direction))
                conn.commit(); conn.close()
            except: pass

        if now-last_image_save>=IMAGE_INTV and current_event_id:
            try:
                ts=datetime.now().strftime("%Y%m%d_%H%M%S")
                fn=f"ev{current_event_id}_{ts}.jpg"
                fp=os.path.join("static","captured_images",fn)
                cv2.imwrite(fp,display)
                conn=get_db()
                conn.execute("INSERT INTO saved_images VALUES (NULL,?,?,?,?,?)",
                    (current_event_id,sid,
                     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),fn,count))
                conn.commit(); conn.close()
                last_image_save=now
                print(f"[HUMIX] Image saved: {fn}")
            except Exception as e: print(f"[HUMIX] Img save: {e}")

        h_f=display.shape[0]
        cv2.rectangle(display,(0,0),(310,120),(0,0,0),-1)
        cv2.putText(display,f"People : {count}",(10,30),cv2.FONT_HERSHEY_SIMPLEX,0.75,(255,255,255),2)
        cv2.putText(display,f"Threshold: {current_threshold}",(10,58),cv2.FONT_HERSHEY_SIMPLEX,0.55,(180,180,180),1)
        cv2.putText(display,datetime.now().strftime("%H:%M:%S"),(10,84),cv2.FONT_HERSHEY_SIMPLEX,0.55,(140,140,140),1)
        cv2.putText(display,flow_direction,(10,110),cv2.FONT_HERSHEY_SIMPLEX,0.45,(200,200,0),1)
        if alert_active:
            cv2.rectangle(display,(0,0),(display.shape[1]-1,h_f-1),(0,0,230),5)
            cv2.putText(display,"ALERT: CROWD LIMIT EXCEEDED",
                (10,h_f-15),cv2.FONT_HERSHEY_SIMPLEX,0.65,(0,0,255),2)

        with frame_lock: latest_frame=display.copy()
        print(f"[HUMIX] Count:{count} | Flow:{flow_direction}")
        last_capture=now

# ── CAMERA MANAGEMENT ───────────────────────────────────
def start_camera(source,url="",event_id=None,threshold=10):
    global camera,camera_active,camera_source,esp32_url
    global capture_thread,session_data,current_event_id,current_threshold,prev_gray
    camera_source=source; esp32_url=url
    current_event_id=event_id; current_threshold=threshold; prev_gray=None

    if source=="webcam":
        camera=cv2.VideoCapture(0)
        camera.set(cv2.CAP_PROP_FRAME_WIDTH,640)
        camera.set(cv2.CAP_PROP_FRAME_HEIGHT,480)
        if not camera.isOpened(): return False
        for _ in range(20): camera.read()

    load_model()
    sid=str(uuid.uuid4())[:8]
    camera_active=True
    session_data.update({"start_time":datetime.now(),"total_logs":0,
                         "alert_events":0,"max_crowd":0,"session_id":sid})
    if event_id:
        try:
            conn=get_db()
            conn.execute("INSERT INTO mon_sessions VALUES (?,?,?,NULL,0,0,0,?)",
                (sid,event_id,datetime.now().strftime("%Y-%m-%d %H:%M:%S"),threshold))
            conn.execute("UPDATE events SET status='active' WHERE id=?",(event_id,))
            conn.commit(); conn.close()
        except: pass

    capture_thread=threading.Thread(target=capture_worker,daemon=True)
    capture_thread.start()
    return True

def stop_camera():
    global camera,camera_active
    sid=session_data.get("session_id")
    if sid and current_event_id:
        try:
            conn=get_db()
            conn.execute(
                "UPDATE mon_sessions SET end_time=?,max_crowd=?,total_alerts=?,total_logs=? WHERE id=?",
                (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                 session_data["max_crowd"],session_data["alert_events"],
                 session_data["total_logs"],sid))
            conn.execute("UPDATE events SET status='completed' WHERE id=?",(current_event_id,))
            conn.commit(); conn.close()
        except: pass
    send_led("G")
    camera_active=False
    if camera: camera.release(); camera=None

# ── AUTH DECORATORS ─────────────────────────────────────
def admin_req(f):
    @wraps(f)
    def d(*a,**k):
        if not session.get("admin"): return redirect(url_for("admin_login_page"))
        return f(*a,**k)
    return d

def event_req(f):
    @wraps(f)
    def d(*a,**k):
        if not session.get("event_id"): return redirect(url_for("event_login_page"))
        return f(*a,**k)
    return d

def api_auth(f):
    @wraps(f)
    def d(*a,**k):
        if not session.get("admin") and not session.get("event_id"):
            return jsonify({"error":"Unauthorized"}),401
        return f(*a,**k)
    return d

def api_admin(f):
    @wraps(f)
    def d(*a,**k):
        if not session.get("admin"): return jsonify({"error":"Admin only"}),403
        return f(*a,**k)
    return d

# ── ROUTES ──────────────────────────────────────────────
@app.route("/")
def landing():
    return render_template("landing.html")

@app.route("/register", methods=["GET","POST"])
def register():
    if request.method=="POST":
        d=request.form
        try:
            conn=get_db()
            conn.execute("""INSERT INTO events
                (event_name,date,start_time,end_time,venue,estimated_crowd,
                 organizer_name,email,password,status,created_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                (d.get("event_name"),d.get("date"),d.get("start_time"),
                 d.get("end_time"),d.get("venue"),int(d.get("estimated_crowd",0)),
                 d.get("organizer_name"),d.get("email"),hp(d.get("password","")),
                 "registered",datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit(); conn.close()
            return redirect(url_for("event_login_page")+"?registered=1")
        except sqlite3.IntegrityError:
            return render_template("register.html",error="Email already registered")
        except Exception as e:
            return render_template("register.html",error=str(e))
    return render_template("register.html")

@app.route("/login", methods=["GET","POST"])
def event_login_page():
    registered=request.args.get("registered")
    if request.method=="POST":
        conn=get_db()
        ev=conn.execute("SELECT * FROM events WHERE email=? AND password=?",
            (request.form.get("email"),hp(request.form.get("password","")))).fetchone()
        conn.close()
        if ev:
            session["event_id"]=ev["id"]
            session["event_name"]=ev["event_name"]
            return redirect(url_for("event_dashboard"))
        return render_template("event_login.html",error="Invalid credentials")
    return render_template("event_login.html",registered=registered)

@app.route("/admin")
def admin_login_page():
    if session.get("admin"): return redirect(url_for("admin_events"))
    return render_template("admin_login.html")

@app.route("/admin/login", methods=["POST"])
def admin_login():
    if (request.form.get("username")==ADMIN_USER and
        request.form.get("password")==ADMIN_PASS):
        session["admin"]=True
        return redirect(url_for("admin_events"))
    return render_template("admin_login.html",error="Invalid admin credentials")

@app.route("/admin/events")
@admin_req
def admin_events():
    conn=get_db()
    events=conn.execute("SELECT * FROM events ORDER BY created_at DESC").fetchall()
    conn.close()
    return render_template("admin_events.html",events=events)

@app.route("/admin/select/<int:eid>")
@admin_req
def admin_select(eid):
    conn=get_db()
    ev=conn.execute("SELECT * FROM events WHERE id=?",(eid,)).fetchone()
    conn.close()
    if ev:
        session["sel_eid"]=eid
        session["sel_ename"]=ev["event_name"]
        session["sel_thresh"]=ev["estimated_crowd"]
        return redirect(url_for("admin_dashboard"))
    return redirect(url_for("admin_events"))

@app.route("/admin/dashboard")
@admin_req
def admin_dashboard():
    if not session.get("sel_eid"): return redirect(url_for("admin_events"))
    conn=get_db()
    ev=conn.execute("SELECT * FROM events WHERE id=?",(session["sel_eid"],)).fetchone()
    conn.close()
    return render_template("admin_dashboard.html",event=ev,camera_active=camera_active)

@app.route("/event/dashboard")
@event_req
def event_dashboard():
    conn=get_db()
    ev=conn.execute("SELECT * FROM events WHERE id=?",(session["event_id"],)).fetchone()
    conn.close()
    return render_template("event_dashboard.html",event=ev)

@app.route("/logout")
def logout():
    session.clear(); return redirect(url_for("landing"))

@app.route("/admin/logout")
def admin_logout():
    session.clear(); stop_camera(); return redirect(url_for("landing"))

# ── API ─────────────────────────────────────────────────
@app.route("/latest_image")
@api_auth
def latest_image():
    if not camera_active or latest_frame is None: return "No image",404
    for _ in range(50):
        with frame_lock: f=latest_frame
        if f is not None: break
        time.sleep(0.1)
    else: return "Not ready",404
    with frame_lock:
        ok,buf=cv2.imencode(".jpg",latest_frame,[cv2.IMWRITE_JPEG_QUALITY,85])
    if not ok: return "Error",500
    return buf.tobytes(),200,{"Content-Type":"image/jpeg","Cache-Control":"no-store"}

@app.route("/video_feed")
@api_auth
def video_feed():
    if not camera_active: return "Not active",404
    return Response(generate_stream(),mimetype="multipart/x-mixed-replace; boundary=frame")

@app.route("/api/camera/start",methods=["POST"])
@api_admin
def api_start():
    d=request.json or {}
    if camera_active: return jsonify({"success":False,"message":"Already running"})
    ok=start_camera(d.get("source","webcam"),d.get("url",""),
        event_id=session.get("sel_eid"),
        threshold=session.get("sel_thresh",10))
    return jsonify({"success":ok,"message":"Started" if ok else "Failed"})

@app.route("/api/camera/stop",methods=["POST"])
@api_admin
def api_stop():
    stop_camera()
    return jsonify({"success":True})

@app.route("/api/status")
@api_auth
def api_status():
    return jsonify({
        "camera_active":camera_active,"camera_source":camera_source,
        "current_count":current_count,"alert_active":alert_active,
        "threshold":current_threshold,"flow_direction":flow_direction,
        "timestamp":datetime.now().strftime("%H:%M:%S"),
    })

@app.route("/api/session-stats")
@api_auth
def api_stats():
    if not session_data["start_time"]:
        return jsonify({"session_start":"N/A","session_duration":0,
                        "total_logs":0,"alert_events":0,"max_crowd":0})
    dur=int((datetime.now()-session_data["start_time"]).total_seconds())
    return jsonify({
        "session_start":session_data["start_time"].strftime("%H:%M:%S"),
        "session_duration":dur,"total_logs":session_data["total_logs"],
        "alert_events":session_data["alert_events"],"max_crowd":session_data["max_crowd"],
    })

@app.route("/api/graph-data")
@api_auth
def api_graph():
    eid=session.get("sel_eid") or session.get("event_id")
    sid=session_data.get("session_id")
    try:
        conn=get_db()
        rows=conn.execute(
            "SELECT timestamp,count,alert FROM crowd_logs WHERE event_id=? AND session_id=? ORDER BY id DESC LIMIT 30",
            (eid,sid)).fetchall()
        conn.close()
        rows=list(reversed(rows))
        return jsonify({"labels":[r["timestamp"] for r in rows],
                        "counts":[r["count"] for r in rows],
                        "alerts":[r["alert"] for r in rows],
                        "threshold":current_threshold})
    except:
        return jsonify({"labels":[],"counts":[],"alerts":[],"threshold":current_threshold})

@app.route("/api/settings/threshold",methods=["POST"])
@api_admin
def api_thresh():
    global current_threshold
    v=(request.json or {}).get("threshold")
    if v and isinstance(v,int) and v>0:
        current_threshold=v
        session["sel_thresh"]=v
        return jsonify({"success":True,"message":f"Threshold set to {v}"})
    return jsonify({"success":False,"message":"Invalid"})

@app.route("/api/settings/confidence",methods=["POST"])
@api_admin
def api_conf():
    return jsonify({"success":True})

@app.route("/api/arduino/connect",methods=["POST"])
@api_admin
def api_arduino():
    port=(request.json or {}).get("port","COM3")
    ok=connect_arduino(port)
    return jsonify({"success":ok,"message":f"Connected on {port}" if ok else "Failed"})

@app.route("/api/report")
@api_auth
def api_report():
    try:
        import openpyxl
        from openpyxl.styles import Font,PatternFill,Alignment
        from openpyxl.chart import LineChart,Reference
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImg

        eid=session.get("sel_eid") or session.get("event_id")
        sid=session_data.get("session_id")
        conn=get_db()
        ev  =conn.execute("SELECT * FROM events WHERE id=?",(eid,)).fetchone()
        logs=conn.execute(
            "SELECT timestamp,count,threshold,alert,flow_direction FROM crowd_logs WHERE event_id=? AND session_id=? ORDER BY id",
            (eid,sid)).fetchall()
        imgs=conn.execute(
            "SELECT timestamp,filename,count FROM saved_images WHERE event_id=? AND session_id=? ORDER BY id LIMIT 3",
            (eid,sid)).fetchall()
        conn.close()

        wb=openpyxl.Workbook()

        # Sheet 1 – Summary
        ws=wb.active; ws.title="Event Summary"
        ws.merge_cells("A1:E1")
        ws["A1"]="HUMIX — Smart Human Density & Movement Monitoring System"
        ws["A1"].font=Font(size=14,bold=True,color="FFFFFF")
        ws["A1"].fill=PatternFill("solid",fgColor="0c0e14")
        ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[1].height=40

        ws.merge_cells("A2:E2")
        ws["A2"]=f"Event Report: {ev['event_name'] if ev else 'N/A'}"
        ws["A2"].font=Font(size=12,bold=True,color="22c55e")
        ws["A2"].fill=PatternFill("solid",fgColor="13161f")
        ws["A2"].alignment=Alignment(horizontal="center")
        ws.row_dimensions[2].height=28

        info=[
            ("",""),("EVENT DETAILS",""),
            ("Event Name",   ev["event_name"]      if ev else ""),
            ("Date",         ev["date"]             if ev else ""),
            ("Start Time",   ev["start_time"]       if ev else ""),
            ("End Time",     ev["end_time"]         if ev else ""),
            ("Venue",        ev["venue"]            if ev else ""),
            ("Organizer",    ev["organizer_name"]   if ev else ""),
            ("Est. Crowd",   ev["estimated_crowd"]  if ev else ""),
            ("",""),("MONITORING SUMMARY",""),
            ("Session Start", session_data["start_time"].strftime("%Y-%m-%d %H:%M:%S") if session_data["start_time"] else "N/A"),
            ("Peak Count",    session_data["max_crowd"]),
            ("Alert Events",  session_data["alert_events"]),
            ("Total Logs",    session_data["total_logs"]),
            ("Threshold",     current_threshold),
            ("Alert Status",  "ALERTS TRIGGERED" if session_data["alert_events"]>0 else "NO ALERTS"),
        ]
        for i,(lbl,val) in enumerate(info,3):
            if lbl in ("EVENT DETAILS","MONITORING SUMMARY"):
                ws.merge_cells(f"A{i}:E{i}")
                ws[f"A{i}"]=lbl
                ws[f"A{i}"].font=Font(bold=True,color="FFFFFF",size=10)
                ws[f"A{i}"].fill=PatternFill("solid",fgColor="22c55e")
            elif lbl:
                ws[f"A{i}"]=lbl; ws[f"B{i}"]=str(val)
                ws[f"A{i}"].font=Font(bold=True)
        ws.column_dimensions["A"].width=22
        ws.column_dimensions["B"].width=35

        # Sheet 2 – Data Log
        ws2=wb.create_sheet("Crowd Data Log")
        hdrs=["Time","Count","Threshold","Alert","Flow Direction"]
        gf=PatternFill("solid",fgColor="22c55e")
        rf=PatternFill("solid",fgColor="FEE2E2")
        for col,h in enumerate(hdrs,1):
            c=ws2.cell(1,col,h)
            c.font=Font(bold=True,color="FFFFFF"); c.fill=gf
            c.alignment=Alignment(horizontal="center")
            ws2.column_dimensions[get_column_letter(col)].width=18
        for rn,row in enumerate(logs,2):
            ws2.cell(rn,1,row["timestamp"]); ws2.cell(rn,2,row["count"])
            ws2.cell(rn,3,row["threshold"])
            ws2.cell(rn,4,"YES" if row["alert"] else "No")
            ws2.cell(rn,5,row["flow_direction"])
            if row["alert"]:
                for col in range(1,6): ws2.cell(rn,col).fill=rf

        # Sheet 3 – Graph
        ws3=wb.create_sheet("Graph")
        ws3["A1"],ws3["B1"],ws3["C1"]="Time","Count","Threshold"
        for i,row in enumerate(logs,2):
            ws3.cell(i,1,row["timestamp"])
            ws3.cell(i,2,row["count"])
            ws3.cell(i,3,row["threshold"])
        if len(logs)>1:
            ch=LineChart()
            ch.title="Crowd Count Over Time"
            ch.style=10; ch.width=28; ch.height=16
            ch.y_axis.title="People"; ch.x_axis.title="Time"
            dr=Reference(ws3,min_col=2,max_col=3,min_row=1,max_row=len(logs)+1)
            lr=Reference(ws3,min_col=1,min_row=2,max_row=len(logs)+1)
            ch.add_data(dr,titles_from_data=True); ch.set_categories(lr)
            ws3.add_chart(ch,"E2")

        # Sheet 4 – Captured Images (up to 3)
        if imgs:
            ws4=wb.create_sheet("Captured Images")
            ws4["A1"]="Captured Monitoring Images (every 3 minutes)"
            ws4["A1"].font=Font(size=13,bold=True)
            ws4.row_dimensions[1].height=28
            rp=3
            for img in imgs:
                fp=os.path.join("static","captured_images",img["filename"])
                ws4.cell(rp,1,f"Time: {img['timestamp']}  |  Count: {img['count']}")
                ws4.cell(rp,1).font=Font(bold=True,color="22c55e")
                rp+=1
                if os.path.exists(fp):
                    try:
                        xi=XLImg(fp); xi.width=420; xi.height=290
                        ws4.add_image(xi,f"A{rp}"); rp+=17
                    except:
                        ws4.cell(rp,1,f"[{img['filename']}]"); rp+=2
                else:
                    rp+=1

        out=io.BytesIO(); wb.save(out); out.seek(0)
        fn=f"HUMIX_{ev['event_name'].replace(' ','_') if ev else 'Report'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,download_name=fn)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error":str(e)}),500

if __name__=="__main__":
    init_db()
    app.run(debug=True,host="0.0.0.0",port=5000)
'''

# ── LANDING.HTML ──────────────────────────────────────────
files[f"{BASE}/templates/landing.html"] = """\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
  <title>HUMIX — Smart Monitoring</title>
  <link rel="stylesheet" href="{{ url_for('static',filename='css/style.css') }}">
</head>
<body class="landing-page">
  <div class="landing-wrap">
    <div class="landing-glow"></div>
    <div class="landing-content">
      <div class="humix-logo">📡</div>
      <h1 class="humix-title">HUMIX</h1>
      <p class="humix-sub">Smart Human Density &amp; Movement Monitoring System</p>
      <p class="humix-desc">
        Real-time crowd monitoring powered by AI detection.<br>
        Register your event and monitor crowd safety with precision.
      </p>
      <div class="landing-btns">
        <a href="/register" class="btn btn-landing-primary">
          <span>📋</span> Register Event
        </a>
        <a href="/admin" class="btn btn-landing-admin">
          <span>🔐</span> Admin Login
        </a>
      </div>
      <div class="landing-features">
        <div class="feat-item"><span>🎯</span><p>AI Detection</p></div>
        <div class="feat-item"><span>📊</span><p>Live Graph</p></div>
        <div class="feat-item"><span>🚨</span><p>Instant Alerts</p></div>
        <div class="feat-item"><span>📥</span><p>Auto Reports</p></div>
      </div>
    </div>
  </div>
</body>
</html>
"""

# ── REGISTER.HTML ─────────────────────────────────────────
files[f"{BASE}/templates/register.html"] = """\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
  <title>HUMIX — Register Event</title>
  <link rel="stylesheet" href="{{ url_for('static',filename='css/style.css') }}">
</head>
<body class="login-page">
  <div class="login-wrap" style="max-width:560px">
    <div class="login-brand">
      <div class="brand-icon">📡</div>
      <h1>HUMIX</h1>
      <p>Register Your Event</p>
    </div>
    <div class="login-card">
      <h2>Event Registration</h2>
      {% if error %}
      <div class="err-msg" style="display:block;margin-bottom:16px">{{ error }}</div>
      {% endif %}
      <form method="POST" action="/register">
        <div class="form-row">
          <div class="field">
            <label>Event Name *</label>
            <input type="text" name="event_name" placeholder="e.g. Tech Fest 2025" required>
          </div>
          <div class="field">
            <label>Venue / Location *</label>
            <input type="text" name="venue" placeholder="e.g. Main Hall, City Stadium" required>
          </div>
        </div>
        <div class="form-row">
          <div class="field">
            <label>Event Date *</label>
            <input type="date" name="date" required>
          </div>
          <div class="field">
            <label>Estimated Crowd *</label>
            <input type="number" name="estimated_crowd" placeholder="e.g. 500" min="1" required>
          </div>
        </div>
        <div class="form-row">
          <div class="field">
            <label>Start Time *</label>
            <input type="time" name="start_time" required>
          </div>
          <div class="field">
            <label>End Time *</label>
            <input type="time" name="end_time" required>
          </div>
        </div>
        <div class="form-row">
          <div class="field">
            <label>Organizer Name *</label>
            <input type="text" name="organizer_name" placeholder="Your full name" required>
          </div>
          <div class="field">
            <label>Email *</label>
            <input type="email" name="email" placeholder="you@email.com" required>
          </div>
        </div>
        <div class="field">
          <label>Password *</label>
          <input type="password" name="password" placeholder="Create a password" required>
        </div>
        <button type="submit" class="btn btn-primary btn-full" style="margin-top:8px">
          ✅ Register Event
        </button>
      </form>
      <div style="text-align:center;margin-top:20px;font-size:.88rem;color:#64748b">
        Already registered? <a href="/login" style="color:#22c55e;font-weight:700">Login here</a>
        &nbsp;|&nbsp;
        <a href="/" style="color:#64748b">← Back</a>
      </div>
    </div>
  </div>
</body>
</html>
"""

# ── EVENT_LOGIN.HTML ──────────────────────────────────────
files[f"{BASE}/templates/event_login.html"] = """\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
  <title>HUMIX — Event Login</title>
  <link rel="stylesheet" href="{{ url_for('static',filename='css/style.css') }}">
</head>
<body class="login-page">
  <div class="login-wrap">
    <div class="login-brand">
      <div class="brand-icon">📡</div>
      <h1>HUMIX</h1>
      <p>Event Organizer Login</p>
    </div>
    <div class="login-card">
      <h2>Sign In</h2>
      {% if registered %}
      <div style="background:rgba(34,197,94,.1);color:#22c55e;border:1px solid rgba(34,197,94,.2);
           padding:12px;border-radius:8px;margin-bottom:16px;text-align:center;font-weight:700">
        ✅ Event registered successfully! Please login.
      </div>
      {% endif %}
      {% if error %}
      <div class="err-msg" style="display:block;margin-bottom:16px">{{ error }}</div>
      {% endif %}
      <form method="POST" action="/login">
        <div class="field">
          <label>Email</label>
          <input type="email" name="email" placeholder="your@email.com" required autofocus>
        </div>
        <div class="field">
          <label>Password</label>
          <input type="password" name="password" placeholder="Your password" required>
        </div>
        <button type="submit" class="btn btn-primary btn-full">Sign In</button>
      </form>
      <div style="text-align:center;margin-top:20px;font-size:.88rem;color:#64748b">
        New event? <a href="/register" style="color:#22c55e;font-weight:700">Register here</a>
        &nbsp;|&nbsp;
        <a href="/" style="color:#64748b">← Back</a>
      </div>
    </div>
  </div>
</body>
</html>
"""

# ── ADMIN_LOGIN.HTML ──────────────────────────────────────
files[f"{BASE}/templates/admin_login.html"] = """\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
  <title>HUMIX — Admin Login</title>
  <link rel="stylesheet" href="{{ url_for('static',filename='css/style.css') }}">
</head>
<body class="login-page">
  <div class="login-wrap">
    <div class="login-brand">
      <div class="brand-icon">🔐</div>
      <h1>Admin</h1>
      <p>HUMIX System Administrator</p>
    </div>
    <div class="login-card">
      <h2>Admin Login</h2>
      {% if error %}
      <div class="err-msg" style="display:block;margin-bottom:16px">{{ error }}</div>
      {% endif %}
      <form method="POST" action="/admin/login">
        <div class="field">
          <label>Username</label>
          <input type="text" name="username" placeholder="admin" required autofocus>
        </div>
        <div class="field">
          <label>Password</label>
          <input type="password" name="password" placeholder="Password" required>
        </div>
        <button type="submit" class="btn btn-primary btn-full">Login as Admin</button>
      </form>
      <div style="text-align:center;margin-top:16px">
        <a href="/" style="color:#64748b;font-size:.88rem">← Back to Home</a>
      </div>
    </div>
  </div>
</body>
</html>
"""

# ── ADMIN_EVENTS.HTML ─────────────────────────────────────
files[f"{BASE}/templates/admin_events.html"] = """\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
  <title>HUMIX — Admin Events</title>
  <link rel="stylesheet" href="{{ url_for('static',filename='css/style.css') }}">
</head>
<body class="dash-page">
  <header class="topbar">
    <div class="topbar-left">
      <span class="topbar-icon">📡</span>
      <span class="topbar-title">HUMIX Admin</span>
    </div>
    <div class="topbar-right">
      <span class="role-pill role-admin">ADMIN</span>
      <a href="/admin/logout" class="btn btn-logout">Logout</a>
    </div>
  </header>

  <div style="max-width:1200px;margin:0 auto;padding:28px 20px">
    <div style="margin-bottom:24px">
      <h2 style="font-size:1.5rem;font-weight:800;color:var(--text)">
        📋 Registered Events
      </h2>
      <p style="color:var(--muted);font-size:.9rem;margin-top:4px">
        Select an event to begin monitoring
      </p>
    </div>

    {% if events %}
    <div class="events-grid">
      {% for ev in events %}
      <div class="event-card">
        <div class="event-card-top">
          <div class="event-name">{{ ev.event_name }}</div>
          <span class="ev-status ev-{{ ev.status }}">{{ ev.status|upper }}</span>
        </div>
        <div class="event-details">
          <div class="ev-row"><span>📅</span><span>{{ ev.date }}</span></div>
          <div class="ev-row"><span>🕐</span><span>{{ ev.start_time }} – {{ ev.end_time }}</span></div>
          <div class="ev-row"><span>📍</span><span>{{ ev.venue }}</span></div>
          <div class="ev-row"><span>👤</span><span>{{ ev.organizer_name }}</span></div>
          <div class="ev-row"><span>👥</span><span>Est. {{ ev.estimated_crowd }} people</span></div>
        </div>
        <a href="/admin/select/{{ ev.id }}" class="btn btn-primary btn-full" style="margin-top:16px">
          🎥 Monitor This Event
        </a>
      </div>
      {% endfor %}
    </div>
    {% else %}
    <div class="no-events">
      <div style="font-size:4rem;margin-bottom:16px">📭</div>
      <p>No events registered yet</p>
      <small>Events will appear here once organizers register</small>
    </div>
    {% endif %}
  </div>
</body>
</html>
"""

# ── ADMIN_DASHBOARD.HTML ──────────────────────────────────
files[f"{BASE}/templates/admin_dashboard.html"] = """\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
  <title>HUMIX — Admin Dashboard</title>
  <link rel="stylesheet" href="{{ url_for('static',filename='css/style.css') }}">
  <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
</head>
<body class="dash-page">
  <header class="topbar">
    <div class="topbar-left">
      <span class="topbar-icon">📡</span>
      <span class="topbar-title">HUMIX</span>
      <span class="event-badge">{{ event.event_name }}</span>
    </div>
    <div class="topbar-right">
      <button id="themeToggle" class="btn btn-theme">🌙</button>
      <a href="/api/report" class="btn btn-report" target="_blank">📥 Report</a>
      <a href="/admin/events" class="btn btn-outline">← Events</a>
      <a href="/admin/logout" class="btn btn-logout">Logout</a>
    </div>
  </header>

  <div id="alertBanner" class="alert-banner" style="display:none">
    🚨 &nbsp; CROWD LIMIT EXCEEDED — Take action immediately!
  </div>

  <div class="dash-layout">
    <aside class="left-panel">

      <!-- Camera Controls -->
      <div class="card">
        <div class="card-title">📷 Camera Controls</div>
        <div class="field">
          <label>Camera Source</label>
          <select id="cameraSource">
            <option value="webcam">💻 Laptop Webcam</option>
            <option value="esp32">📡 ESP32-CAM</option>
          </select>
        </div>
        <div class="field" id="esp32Field" style="display:none">
          <label>ESP32 URL</label>
          <input type="text" id="esp32Url" placeholder="http://192.168.x.x/capture">
        </div>
        <div class="cam-btns">
          <button id="startBtn" class="btn btn-green">▶ Start</button>
          <button id="stopBtn"  class="btn btn-red" disabled>⏹ Stop</button>
        </div>
        <div class="cam-status-row">
          <span class="status-label">Status</span>
          <span id="camStatus" class="status-pill pill-offline">Offline</span>
          <span id="srcTag" class="source-tag" style="display:none"></span>
        </div>
      </div>

      <!-- Arduino -->
      <div class="card">
        <div class="card-title">💡 Arduino LED Alert</div>
        <div class="field">
          <label>COM Port</label>
          <input type="text" id="arduinoPort" value="COM3" placeholder="e.g. COM3">
        </div>
        <button id="arduinoBtn" class="btn btn-primary btn-full">Connect Arduino</button>
        <div id="arduinoStatus" class="led-status-row" style="margin-top:10px">
          <span class="led green-led" id="ledG" title="Normal">●</span>
          <span class="led yellow-led" id="ledY" title="Warning 75%">●</span>
          <span class="led red-led"   id="ledR" title="Alert">●</span>
          <span style="font-size:.75rem;color:var(--muted);margin-left:6px">LED Status</span>
        </div>
      </div>

      <!-- Settings -->
      <div class="card">
        <div class="card-title">⚙️ Settings</div>
        <div class="field">
          <label>Crowd Threshold</label>
          <input type="number" id="thresholdInput"
                 value="{{ event.estimated_crowd }}" min="1" max="5000">
        </div>
        <div class="field">
          <label>Confidence (<span id="confLabel">0.50</span>)</label>
          <input type="range" id="confInput" min="10" max="90" value="50" step="5">
        </div>
        <button id="saveBtn" class="btn btn-primary btn-full">Save Settings</button>
      </div>

      <!-- Live Stats -->
      <div class="card">
        <div class="card-title">📊 Live Status</div>
        <div class="stat-grid">
          <div class="stat-box">
            <div class="stat-val" id="countVal">0</div>
            <div class="stat-lbl">Detected</div>
          </div>
          <div class="stat-box">
            <div class="stat-val" id="threshVal">—</div>
            <div class="stat-lbl">Threshold</div>
          </div>
          <div class="stat-box stat-wide">
            <div class="stat-val stat-time" id="timeVal">--:--:--</div>
            <div class="stat-lbl">Last Update</div>
          </div>
        </div>
        <div class="flow-box">
          <span class="flow-label">Crowd Flow</span>
          <span class="flow-val" id="flowVal">—</span>
        </div>
      </div>

      <!-- Session Stats -->
      <div class="card" id="sessionCard" style="display:none">
        <div class="card-title">📈 Session Statistics</div>
        <div class="stat-rows">
          <div class="stat-row"><span>Started</span>      <span id="sesStart">—</span></div>
          <div class="stat-row"><span>Duration</span>     <span id="sesDur">0s</span></div>
          <div class="stat-row"><span>Snapshots</span>    <span id="sesLogs">0</span></div>
          <div class="stat-row"><span>Alert Events</span> <span id="sesAlerts">0</span></div>
          <div class="stat-row"><span>Peak Count</span>   <span id="sesPeak">0</span></div>
        </div>
      </div>

    </aside>

    <main class="right-panel">

      <!-- Captured Image -->
      <div class="card card-feed">
        <div class="card-title">🖼️ Captured Image
          <span class="capture-note">Preprocessed · Every 2 seconds</span>
        </div>
        <div class="feed-box">
          <img id="feedImg" src="" style="display:none">
          <div id="feedPh" class="feed-placeholder">
            <div class="ph-icon">📷</div>
            <p>Camera not active</p>
            <small>Select source and press Start</small>
          </div>
        </div>
      </div>

      <!-- Graph -->
      <div class="card card-graph">
        <div class="card-title">📉 Crowd Count Over Time
          <span class="capture-note">Live</span>
        </div>
        <div class="graph-container">
          <canvas id="crowdChart"></canvas>
          <div class="graph-placeholder" id="graphPh">
            <p>Graph appears when camera starts</p>
          </div>
        </div>
      </div>

    </main>
  </div>

  <script>const USER_ROLE="admin";</script>
  <script src="{{ url_for('static',filename='js/admin_dashboard.js') }}"></script>
</body>
</html>
"""

# ── EVENT_DASHBOARD.HTML ──────────────────────────────────
files[f"{BASE}/templates/event_dashboard.html"] = """\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
  <title>HUMIX — {{ event.event_name }}</title>
  <link rel="stylesheet" href="{{ url_for('static',filename='css/style.css') }}">
  <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  <audio id="alertSound" loop>
    <source src="https://assets.mixkit.co/active_storage/sfx/2869/2869-preview.mp3" type="audio/mpeg">
  </audio>
</head>
<body class="dash-page">
  <header class="topbar">
    <div class="topbar-left">
      <span class="topbar-icon">📡</span>
      <span class="topbar-title">HUMIX</span>
      <span class="event-badge">{{ event.event_name }}</span>
    </div>
    <div class="topbar-right">
      <button id="themeToggle" class="btn btn-theme">🌙</button>
      <span class="role-pill role-event">EVENT</span>
      <a href="/api/report" class="btn btn-report" target="_blank">📥 Report</a>
      <a href="/logout" class="btn btn-logout">Logout</a>
    </div>
  </header>

  <div id="alertBanner" class="alert-banner" style="display:none">
    🚨 &nbsp; CROWD LIMIT EXCEEDED — Take action immediately!
    <button id="muteBtn" onclick="toggleMute()" class="mute-btn">🔊 Mute</button>
  </div>

  <div class="dash-layout">
    <aside class="left-panel">

      <!-- Event Info -->
      <div class="card">
        <div class="card-title">📋 Event Details</div>
        <div class="stat-rows">
          <div class="stat-row"><span>Event</span>      <span>{{ event.event_name }}</span></div>
          <div class="stat-row"><span>Date</span>       <span>{{ event.date }}</span></div>
          <div class="stat-row"><span>Time</span>       <span>{{ event.start_time }} – {{ event.end_time }}</span></div>
          <div class="stat-row"><span>Venue</span>      <span>{{ event.venue }}</span></div>
          <div class="stat-row"><span>Est. Crowd</span> <span>{{ event.estimated_crowd }}</span></div>
        </div>
      </div>

      <!-- Live Stats -->
      <div class="card">
        <div class="card-title">📊 Live Status</div>
        <div class="stat-grid">
          <div class="stat-box">
            <div class="stat-val" id="countVal">0</div>
            <div class="stat-lbl">Detected</div>
          </div>
          <div class="stat-box">
            <div class="stat-val" id="threshVal">—</div>
            <div class="stat-lbl">Threshold</div>
          </div>
          <div class="stat-box stat-wide">
            <div class="stat-val stat-time" id="timeVal">--:--:--</div>
            <div class="stat-lbl">Last Update</div>
          </div>
        </div>
        <div class="flow-box">
          <span class="flow-label">Crowd Flow</span>
          <span class="flow-val" id="flowVal">—</span>
        </div>
      </div>

      <!-- Session Stats -->
      <div class="card" id="sessionCard" style="display:none">
        <div class="card-title">📈 Session Statistics</div>
        <div class="stat-rows">
          <div class="stat-row"><span>Started</span>      <span id="sesStart">—</span></div>
          <div class="stat-row"><span>Duration</span>     <span id="sesDur">0s</span></div>
          <div class="stat-row"><span>Snapshots</span>    <span id="sesLogs">0</span></div>
          <div class="stat-row"><span>Alert Events</span> <span id="sesAlerts">0</span></div>
          <div class="stat-row"><span>Peak Count</span>   <span id="sesPeak">0</span></div>
        </div>
      </div>

    </aside>

    <main class="right-panel">

      <!-- Live Video -->
      <div class="card card-feed">
        <div class="card-title">🎥 Live Video Feed
          <span class="capture-note" id="liveTag">Waiting for camera...</span>
        </div>
        <div class="feed-box">
          <img id="liveStream" src="" style="display:none;width:100%">
          <div id="feedPh" class="feed-placeholder">
            <div class="ph-icon">🎥</div>
            <p>Camera not active</p>
            <small>Waiting for admin to start camera</small>
          </div>
        </div>
      </div>

      <!-- Graph -->
      <div class="card card-graph">
        <div class="card-title">📉 Crowd Count Over Time
          <span class="capture-note">Live</span>
        </div>
        <div class="graph-container">
          <canvas id="crowdChart"></canvas>
          <div class="graph-placeholder" id="graphPh">
            <p>Graph appears when camera starts</p>
          </div>
        </div>
      </div>

    </main>
  </div>

  <script>const USER_ROLE="event";</script>
  <script src="{{ url_for('static',filename='js/event_dashboard.js') }}"></script>
</body>
</html>
"""

# ── STYLE.CSS ─────────────────────────────────────────────
files[f"{BASE}/static/css/style.css"] = """\
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#0c0e14;--surface:#13161f;--card:#1a1e2a;--border:rgba(255,255,255,0.07);
  --green:#22c55e;--red:#ef4444;--blue:#3b82f6;--cyan:#06b6d4;--yellow:#eab308;
  --text:#e2e8f0;--muted:#64748b;--radius:16px;
}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
     background:var(--bg);color:var(--text);min-height:100vh;}

/* LIGHT MODE */
body.light{
  --bg:#f1f5f9;--surface:#ffffff;--card:#ffffff;
  --border:rgba(0,0,0,0.08);--text:#0f172a;--muted:#64748b;
}
body.light .topbar{box-shadow:0 1px 4px rgba(0,0,0,.08);}
body.light .card{box-shadow:0 2px 12px rgba(0,0,0,.06);}
body.light .stat-box,.body.light .cam-status-row,body.light .stat-row{background:#f8fafc;}
body.light .stat-val{color:#16a34a;}
body.light .field input,body.light .field select{background:#f8fafc;color:#0f172a;border-color:rgba(0,0,0,.12);}
body.light .field label,body.light .card-title{color:#475569;}
body.light .login-page{background:radial-gradient(ellipse 80% 60% at 50% 40%,rgba(34,197,94,.06) 0%,transparent 70%),#f1f5f9;}
body.light .login-card{box-shadow:0 12px 40px rgba(0,0,0,.1);}
body.light .event-card{background:#fff;box-shadow:0 2px 12px rgba(0,0,0,.06);}
body.light .no-events{color:#475569;}

/* LANDING */
.landing-page{display:flex;align-items:center;justify-content:center;
  min-height:100vh;padding:24px;overflow:hidden;
  background:radial-gradient(ellipse 100% 80% at 50% 30%,rgba(34,197,94,.12) 0%,transparent 60%),var(--bg);}
.landing-wrap{width:100%;max-width:680px;position:relative;text-align:center;}
.landing-glow{position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);
  width:600px;height:400px;background:rgba(34,197,94,.06);
  filter:blur(80px);border-radius:50%;pointer-events:none;}
.landing-content{position:relative;z-index:1;}
.humix-logo{font-size:5rem;margin-bottom:16px;filter:drop-shadow(0 0 20px rgba(34,197,94,.4));}
.humix-title{font-size:4rem;font-weight:900;letter-spacing:-2px;
  background:linear-gradient(135deg,var(--green),var(--cyan));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;
  margin-bottom:12px;}
.humix-sub{font-size:1.1rem;color:#94a3b8;font-weight:600;margin-bottom:10px;}
.humix-desc{font-size:.9rem;color:var(--muted);line-height:1.6;margin-bottom:40px;}
.landing-btns{display:flex;gap:16px;justify-content:center;flex-wrap:wrap;margin-bottom:48px;}
.btn-landing-primary{background:var(--green);color:#000;font-size:1rem;padding:14px 32px;
  border-radius:14px;font-weight:700;display:flex;align-items:center;gap:8px;}
.btn-landing-primary:hover{background:#16a34a;transform:translateY(-2px);
  box-shadow:0 12px 32px rgba(34,197,94,.3);}
.btn-landing-admin{background:rgba(255,255,255,.06);color:var(--text);font-size:1rem;
  padding:14px 32px;border-radius:14px;font-weight:700;
  border:1px solid var(--border);display:flex;align-items:center;gap:8px;}
.btn-landing-admin:hover{background:rgba(255,255,255,.1);transform:translateY(-2px);}
.landing-features{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;}
.feat-item{background:rgba(255,255,255,.03);border:1px solid var(--border);
  border-radius:14px;padding:20px 10px;}
.feat-item span{font-size:1.8rem;display:block;margin-bottom:8px;}
.feat-item p{font-size:.8rem;color:var(--muted);font-weight:600;}
@media(max-width:600px){.landing-features{grid-template-columns:repeat(2,1fr);}
  .humix-title{font-size:2.8rem;}}

/* LOGIN */
.login-page{display:flex;align-items:center;justify-content:center;
  min-height:100vh;padding:24px;
  background:radial-gradient(ellipse 80% 60% at 50% 40%,rgba(34,197,94,.08) 0%,transparent 70%),var(--bg);}
.login-wrap{width:100%;max-width:420px;}
.login-brand{text-align:center;margin-bottom:32px;}
.brand-icon{font-size:3rem;margin-bottom:10px;}
.login-brand h1{font-size:2.2rem;font-weight:900;letter-spacing:-1px;
  background:linear-gradient(135deg,var(--green),var(--cyan));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}
.login-brand p{color:var(--muted);font-size:.88rem;margin-top:6px;}
.login-card{background:var(--card);border:1px solid var(--border);border-radius:24px;
  padding:36px;box-shadow:0 24px 64px rgba(0,0,0,.5);}
.login-card h2{font-size:1.5rem;font-weight:700;margin-bottom:24px;}
.form-row{display:grid;grid-template-columns:1fr 1fr;gap:12px;}
@media(max-width:500px){.form-row{grid-template-columns:1fr;}}

/* FIELDS */
.field{margin-bottom:16px;}
.field label{display:block;margin-bottom:6px;font-size:.82rem;font-weight:600;color:#94a3b8;}
.field input,.field select{width:100%;padding:11px 14px;
  background:rgba(255,255,255,.04);border:1.5px solid var(--border);
  border-radius:10px;color:var(--text);font-size:.95rem;transition:.2s;}
.field input:focus,.field select:focus{outline:none;border-color:var(--green);
  box-shadow:0 0 0 3px rgba(34,197,94,.12);}
.field input[type=range]{accent-color:var(--green);background:transparent;border:none;padding:0;margin-top:4px;}
.field select option{background:var(--card);}

/* BUTTONS */
.btn{padding:10px 20px;border:none;border-radius:10px;font-size:.9rem;font-weight:700;
  cursor:pointer;transition:.2s;display:inline-block;text-decoration:none;}
.btn-primary{background:var(--green);color:#000;}
.btn-primary:hover{background:#16a34a;transform:translateY(-1px);}
.btn-green{background:var(--green);color:#000;flex:1;}
.btn-green:hover:not(:disabled){background:#16a34a;}
.btn-red{background:rgba(239,68,68,.15);color:var(--red);border:1px solid rgba(239,68,68,.25);flex:1;}
.btn-red:hover:not(:disabled){background:var(--red);color:#fff;}
.btn-logout{background:rgba(239,68,68,.1);color:var(--red);border:1px solid rgba(239,68,68,.2);
  font-size:.82rem;padding:7px 16px;}
.btn-logout:hover{background:var(--red);color:#fff;}
.btn-report{background:rgba(59,130,246,.15);color:var(--blue);border:1px solid rgba(59,130,246,.25);
  font-size:.82rem;padding:7px 16px;}
.btn-report:hover{background:var(--blue);color:#fff;}
.btn-outline{background:rgba(255,255,255,.05);color:var(--muted);border:1px solid var(--border);
  font-size:.82rem;padding:7px 16px;}
.btn-outline:hover{background:rgba(255,255,255,.1);color:var(--text);}
.btn-theme{background:rgba(255,255,255,.08);border:1px solid var(--border);
  color:var(--text);padding:7px 12px;font-size:1rem;}
.btn-theme:hover{transform:scale(1.1);}
.btn-full{width:100%;text-align:center;margin-top:6px;}
.btn:disabled{opacity:.35;cursor:not-allowed;}
.err-msg{padding:10px 14px;border-radius:8px;background:rgba(239,68,68,.1);color:var(--red);
  border:1px solid rgba(239,68,68,.2);font-size:.85rem;text-align:center;}

/* TOPBAR */
.topbar{display:flex;justify-content:space-between;align-items:center;
  padding:14px 28px;background:var(--surface);border-bottom:1px solid var(--border);
  position:sticky;top:0;z-index:100;}
.topbar-left{display:flex;align-items:center;gap:10px;}
.topbar-icon{font-size:1.4rem;}
.topbar-title{font-size:1.15rem;font-weight:900;letter-spacing:-.5px;}
.event-badge{background:rgba(34,197,94,.1);color:var(--green);border:1px solid rgba(34,197,94,.2);
  padding:3px 12px;border-radius:20px;font-size:.78rem;font-weight:700;}
.topbar-right{display:flex;align-items:center;gap:8px;}
.role-pill{padding:3px 12px;border-radius:20px;font-size:.7rem;font-weight:700;}
.role-admin{background:var(--green);color:#000;}
.role-event{background:var(--cyan);color:#000;}

/* ALERT */
.alert-banner{background:linear-gradient(90deg,#dc2626,#b91c1c);color:#fff;
  font-weight:700;font-size:.95rem;text-align:center;padding:12px;
  display:flex;align-items:center;justify-content:center;gap:16px;
  animation:flashBanner 1.2s infinite;}
@keyframes flashBanner{0%,100%{opacity:1}50%{opacity:.75}}
.mute-btn{padding:4px 14px;border-radius:8px;border:1px solid rgba(255,255,255,.3);
  background:rgba(255,255,255,.15);color:#fff;font-size:.82rem;font-weight:700;cursor:pointer;}
.mute-btn:hover{background:rgba(255,255,255,.25);}

/* DASHBOARD LAYOUT */
.dash-page{background:var(--bg);}
.dash-layout{display:grid;grid-template-columns:320px 1fr;
  gap:20px;padding:20px;max-width:1700px;margin:0 auto;}
@media(max-width:1100px){.dash-layout{grid-template-columns:1fr;}}

/* CARDS */
.card{background:var(--card);border:1px solid var(--border);
  border-radius:var(--radius);padding:24px;margin-bottom:16px;}
.card:last-child{margin-bottom:0;}
.card-title{font-size:.85rem;font-weight:700;color:#94a3b8;text-transform:uppercase;
  letter-spacing:.8px;margin-bottom:18px;display:flex;align-items:center;gap:6px;flex-wrap:wrap;}
.capture-note{margin-left:auto;font-size:.7rem;font-weight:500;color:var(--green);
  background:rgba(34,197,94,.1);padding:2px 10px;border-radius:20px;
  text-transform:none;letter-spacing:0;}

/* CAMERA CONTROLS */
.cam-btns{display:flex;gap:10px;margin-bottom:14px;}
.cam-status-row{display:flex;align-items:center;gap:10px;padding:10px 14px;
  background:rgba(255,255,255,.03);border-radius:10px;border:1px solid var(--border);}
.status-label{font-size:.82rem;font-weight:600;color:var(--muted);}
.status-pill{padding:4px 12px;border-radius:20px;font-size:.72rem;font-weight:700;}
.pill-online{background:rgba(34,197,94,.15);color:var(--green);border:1px solid rgba(34,197,94,.3);}
.pill-offline{background:rgba(239,68,68,.1);color:var(--red);border:1px solid rgba(239,68,68,.2);}
.source-tag{margin-left:auto;font-size:.75rem;color:var(--cyan);font-weight:600;}

/* ARDUINO LED STATUS */
.led-status-row{display:flex;align-items:center;gap:6px;padding:10px 14px;
  background:rgba(255,255,255,.03);border-radius:10px;border:1px solid var(--border);}
.led{font-size:1.4rem;opacity:.25;transition:opacity .3s;}
.led.active{opacity:1;}
.green-led{color:#22c55e;}
.yellow-led{color:#eab308;}
.red-led{color:#ef4444;}

/* STATS */
.stat-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;}
.stat-wide{grid-column:1/-1;}
.stat-box{background:rgba(255,255,255,.03);border:1px solid var(--border);
  border-radius:12px;padding:18px 14px;text-align:center;}
.stat-val{font-size:2.2rem;font-weight:900;color:var(--green);font-variant-numeric:tabular-nums;}
.stat-time{font-size:1.4rem;}
.stat-lbl{font-size:.7rem;color:var(--muted);text-transform:uppercase;letter-spacing:.6px;margin-top:4px;}
.stat-rows{display:flex;flex-direction:column;gap:8px;}
.stat-row{display:flex;justify-content:space-between;align-items:center;padding:10px 14px;
  background:rgba(255,255,255,.03);border-radius:10px;border:1px solid var(--border);font-size:.85rem;}
.stat-row span:first-child{color:var(--muted);font-weight:600;}
.stat-row span:last-child{color:var(--green);font-weight:700;}
.flow-box{display:flex;justify-content:space-between;align-items:center;margin-top:14px;
  padding:12px 14px;background:rgba(234,179,8,.05);
  border:1px solid rgba(234,179,8,.15);border-radius:10px;}
.flow-label{font-size:.8rem;font-weight:600;color:var(--muted);}
.flow-val{font-size:.9rem;font-weight:700;color:var(--yellow);}

/* FEED & GRAPH */
.card-feed{margin-bottom:16px;}
.right-panel{display:flex;flex-direction:column;}
.feed-box{background:#000;border-radius:12px;overflow:hidden;min-height:380px;
  display:flex;align-items:center;justify-content:center;
  border:1px solid rgba(34,197,94,.15);}
#feedImg{width:100%;height:auto;display:block;}
.feed-placeholder{text-align:center;color:var(--muted);padding:40px;}
.ph-icon{font-size:4rem;opacity:.25;margin-bottom:16px;}
.feed-placeholder p{font-size:1.1rem;color:#94a3b8;font-weight:600;}
.feed-placeholder small{font-size:.85rem;opacity:.6;}
.card-graph{}
.graph-container{position:relative;min-height:260px;display:flex;
  align-items:center;justify-content:center;}
#crowdChart{width:100%!important;}
.graph-placeholder{position:absolute;text-align:center;color:var(--muted);font-size:.9rem;}

/* ADMIN EVENTS PAGE */
.events-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:20px;}
.event-card{background:var(--card);border:1px solid var(--border);border-radius:var(--radius);padding:24px;}
.event-card-top{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:16px;}
.event-name{font-size:1.1rem;font-weight:800;color:var(--text);}
.ev-status{padding:3px 12px;border-radius:20px;font-size:.7rem;font-weight:700;}
.ev-registered{background:rgba(59,130,246,.15);color:var(--blue);}
.ev-active{background:rgba(34,197,94,.15);color:var(--green);}
.ev-completed{background:rgba(100,116,139,.15);color:var(--muted);}
.event-details{display:flex;flex-direction:column;gap:8px;}
.ev-row{display:flex;align-items:center;gap:8px;font-size:.85rem;color:#94a3b8;}
.ev-row span:first-child{font-size:1rem;}
.no-events{text-align:center;padding:80px 40px;color:var(--muted);}
.no-events p{font-size:1.1rem;font-weight:600;color:#94a3b8;margin-bottom:8px;}
"""

# ── ADMIN_DASHBOARD.JS ────────────────────────────────────
files[f"{BASE}/static/js/admin_dashboard.js"] = """\
const startBtn    = document.getElementById('startBtn');
const stopBtn     = document.getElementById('stopBtn');
const camStatus   = document.getElementById('camStatus');
const srcTag      = document.getElementById('srcTag');
const camSrc      = document.getElementById('cameraSource');
const esp32Field  = document.getElementById('esp32Field');
const esp32Url    = document.getElementById('esp32Url');
const threshInput = document.getElementById('thresholdInput');
const confInput   = document.getElementById('confInput');
const confLabel   = document.getElementById('confLabel');
const saveBtn     = document.getElementById('saveBtn');
const feedImg     = document.getElementById('feedImg');
const feedPh      = document.getElementById('feedPh');
const alertBanner = document.getElementById('alertBanner');
const countVal    = document.getElementById('countVal');
const threshVal   = document.getElementById('threshVal');
const timeVal     = document.getElementById('timeVal');
const flowVal     = document.getElementById('flowVal');
const sessionCard = document.getElementById('sessionCard');
const graphPh     = document.getElementById('graphPh');
const ledG        = document.getElementById('ledG');
const ledY        = document.getElementById('ledY');
const ledR        = document.getElementById('ledR');

let statusInt=null, imageInt=null, graphInt=null;
let cameraActive=false, crowdChart=null;

// ── CHART ─────────────────────────────────────────────────
function initChart(){
  const ctx=document.getElementById('crowdChart').getContext('2d');
  crowdChart=new Chart(ctx,{
    type:'line',
    data:{labels:[],datasets:[
      {label:'People Count',data:[],borderColor:'#22c55e',
       backgroundColor:'rgba(34,197,94,.08)',borderWidth:2,
       pointRadius:3,pointBackgroundColor:'#22c55e',tension:.4,fill:true},
      {label:'Threshold',data:[],borderColor:'#ef4444',borderWidth:2,
       borderDash:[6,4],pointRadius:0,tension:0,fill:false}
    ]},
    options:{responsive:true,maintainAspectRatio:true,animation:{duration:400},
      plugins:{legend:{labels:{color:'#94a3b8',font:{size:12}}}},
      scales:{
        x:{ticks:{color:'#64748b',maxTicksLimit:8,maxRotation:0},
           grid:{color:'rgba(255,255,255,.04)'}},
        y:{beginAtZero:true,ticks:{color:'#64748b',stepSize:1},
           grid:{color:'rgba(255,255,255,.06)'}}
      }}
  });
}

async function updateChart(){
  if(!cameraActive||!crowdChart) return;
  try{
    const d=await fetch('/api/graph-data').then(r=>r.json());
    if(!d.labels.length) return;
    if(graphPh) graphPh.style.display='none';
    crowdChart.data.labels=d.labels;
    crowdChart.data.datasets[0].data=d.counts;
    crowdChart.data.datasets[1].data=d.labels.map(()=>d.threshold);
    crowdChart.data.datasets[0].pointBackgroundColor=d.alerts.map(a=>a?'#ef4444':'#22c55e');
    crowdChart.update();
  }catch{}
}

// ── LED DISPLAY ───────────────────────────────────────────
function updateLEDDisplay(count,threshold){
  ledG.classList.remove('active');
  ledY.classList.remove('active');
  ledR.classList.remove('active');
  if(count>=threshold)       ledR.classList.add('active');
  else if(count>=threshold*0.75) ledY.classList.add('active');
  else                       ledG.classList.add('active');
}

// ── SOURCE TOGGLE ─────────────────────────────────────────
camSrc.addEventListener('change',()=>{
  esp32Field.style.display=camSrc.value==='esp32'?'block':'none';
});
confInput.addEventListener('input',()=>{
  confLabel.textContent=(confInput.value/100).toFixed(2);
});

// ── START ─────────────────────────────────────────────────
startBtn.addEventListener('click',async()=>{
  startBtn.disabled=true; startBtn.textContent='Starting…';
  const src=camSrc.value, url=esp32Url?esp32Url.value.trim():'';
  if(src==='esp32'&&!url){
    showToast('Enter ESP32 URL','error');
    startBtn.disabled=false; startBtn.innerHTML='▶ Start'; return;
  }
  try{
    const d=await fetch('/api/camera/start',{method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({source:src,url})}).then(r=>r.json());
    if(d.success){
      cameraActive=true; startBtn.disabled=true; stopBtn.disabled=false;
      camStatus.textContent='Online'; camStatus.className='status-pill pill-online';
      srcTag.textContent=src==='esp32'?'📡 ESP32':'💻 Webcam';
      srcTag.style.display='inline';
      showFeed(); startPolling();
      showToast('Camera started','success');
    }else{ showToast(d.message||'Failed','error'); startBtn.disabled=false; startBtn.innerHTML='▶ Start'; }
  }catch{ showToast('Connection error','error'); startBtn.disabled=false; startBtn.innerHTML='▶ Start'; }
});

// ── STOP ──────────────────────────────────────────────────
stopBtn.addEventListener('click',async()=>{
  stopBtn.disabled=true; stopBtn.textContent='Stopping…';
  try{
    await fetch('/api/camera/stop',{method:'POST'});
    cameraActive=false; startBtn.disabled=false; startBtn.innerHTML='▶ Start';
    stopBtn.disabled=true; stopBtn.innerHTML='⏹ Stop';
    camStatus.textContent='Offline'; camStatus.className='status-pill pill-offline';
    srcTag.style.display='none';
    hideFeed(); stopPolling();
    countVal.textContent='0'; alertBanner.style.display='none';
    if(sessionCard) sessionCard.style.display='none';
    [ledG,ledY,ledR].forEach(l=>l.classList.remove('active'));
    showToast('Camera stopped','success');
  }catch{ showToast('Error','error'); stopBtn.disabled=false; stopBtn.innerHTML='⏹ Stop'; }
});

// ── SETTINGS ──────────────────────────────────────────────
saveBtn.addEventListener('click',async()=>{
  const t=parseInt(threshInput.value), c=parseInt(confInput.value)/100;
  if(t<1||t>5000){showToast('Threshold 1–5000','error');return;}
  try{
    await fetch('/api/settings/threshold',{method:'POST',
      headers:{'Content-Type':'application/json'},body:JSON.stringify({threshold:t})});
    await fetch('/api/settings/confidence',{method:'POST',
      headers:{'Content-Type':'application/json'},body:JSON.stringify({confidence:c})});
    showToast('Settings saved','success');
  }catch{showToast('Failed','error');}
});

// ── ARDUINO ───────────────────────────────────────────────
document.getElementById('arduinoBtn').addEventListener('click',async()=>{
  const port=document.getElementById('arduinoPort').value.trim();
  try{
    const d=await fetch('/api/arduino/connect',{method:'POST',
      headers:{'Content-Type':'application/json'},body:JSON.stringify({port})}).then(r=>r.json());
    showToast(d.message, d.success?'success':'error');
  }catch{showToast('Connection error','error');}
});

// ── POLLING ───────────────────────────────────────────────
function startPolling(){
  pollStatus(); statusInt=setInterval(pollStatus,2000);
  refreshImage(); imageInt=setInterval(refreshImage,2000);
  updateChart(); graphInt=setInterval(updateChart,3000);
}
function stopPolling(){
  clearInterval(statusInt); clearInterval(imageInt); clearInterval(graphInt);
}

async function pollStatus(){
  try{
    const d=await fetch('/api/status').then(r=>r.json());
    countVal.textContent=d.current_count; threshVal.textContent=d.threshold;
    timeVal.textContent=d.timestamp; flowVal.textContent=d.flow_direction||'—';
    alertBanner.style.display=d.alert_active?'block':'none';
    updateLEDDisplay(d.current_count,d.threshold);
    if(d.camera_active&&!cameraActive){cameraActive=true;showFeed();}
    else if(!d.camera_active&&cameraActive){cameraActive=false;hideFeed();}
    if(d.camera_active) pollStats();
  }catch{}
}

async function pollStats(){
  try{
    const d=await fetch('/api/session-stats').then(r=>r.json());
    if(d.session_start!=='N/A'&&sessionCard){
      sessionCard.style.display='block';
      document.getElementById('sesStart').textContent=d.session_start;
      document.getElementById('sesDur').textContent=d.session_duration+'s';
      document.getElementById('sesLogs').textContent=d.total_logs;
      document.getElementById('sesAlerts').textContent=d.alert_events;
      document.getElementById('sesPeak').textContent=d.max_crowd;
    }
  }catch{}
}

function refreshImage(){
  if(!cameraActive) return;
  feedImg.src='/latest_image?t='+Date.now();
  feedImg.onerror=()=>setTimeout(()=>{if(cameraActive)feedImg.src='/latest_image?t='+Date.now();},1000);
}
function showFeed(){
  feedPh.style.display='none'; feedImg.style.display='block';
  setTimeout(()=>{feedImg.src='/latest_image?t='+Date.now();},2200);
}
function hideFeed(){
  feedImg.style.display='none'; feedPh.style.display='flex'; feedImg.src='';
}

// ── THEME ─────────────────────────────────────────────────
const themeToggle=document.getElementById('themeToggle');
const savedTheme=localStorage.getItem('humix_theme')||'dark';
function applyTheme(t){
  if(t==='light'){document.body.classList.add('light');if(themeToggle)themeToggle.textContent='☀️';}
  else{document.body.classList.remove('light');if(themeToggle)themeToggle.textContent='🌙';}
}
applyTheme(savedTheme);
if(themeToggle) themeToggle.addEventListener('click',()=>{
  const nt=document.body.classList.contains('light')?'dark':'light';
  localStorage.setItem('humix_theme',nt); applyTheme(nt);
});

// ── TOAST ─────────────────────────────────────────────────
function showToast(msg,type){
  const t=document.createElement('div');
  t.textContent=msg;
  Object.assign(t.style,{position:'fixed',bottom:'28px',right:'28px',
    padding:'12px 22px',borderRadius:'10px',fontWeight:'700',fontSize:'.88rem',
    color:type==='success'?'#000':'#fff',
    background:type==='success'?'#22c55e':'#ef4444',
    boxShadow:'0 8px 32px rgba(0,0,0,.4)',zIndex:'9999',transition:'opacity .3s'});
  document.body.appendChild(t);
  setTimeout(()=>{t.style.opacity='0';setTimeout(()=>t.remove(),300);},3000);
}

// ── INIT ──────────────────────────────────────────────────
initChart(); pollStatus(); startPolling();
"""

# ── EVENT_DASHBOARD.JS ────────────────────────────────────
files[f"{BASE}/static/js/event_dashboard.js"] = """\
const liveStream  = document.getElementById('liveStream');
const feedPh      = document.getElementById('feedPh');
const alertBanner = document.getElementById('alertBanner');
const countVal    = document.getElementById('countVal');
const threshVal   = document.getElementById('threshVal');
const timeVal     = document.getElementById('timeVal');
const flowVal     = document.getElementById('flowVal');
const sessionCard = document.getElementById('sessionCard');
const graphPh     = document.getElementById('graphPh');
const alertSound  = document.getElementById('alertSound');
const liveTag     = document.getElementById('liveTag');

let statusInt=null, graphInt=null, cameraActive=false;
let crowdChart=null, isMuted=false;

// ── CHART ─────────────────────────────────────────────────
function initChart(){
  const ctx=document.getElementById('crowdChart').getContext('2d');
  crowdChart=new Chart(ctx,{
    type:'line',
    data:{labels:[],datasets:[
      {label:'People Count',data:[],borderColor:'#22c55e',
       backgroundColor:'rgba(34,197,94,.08)',borderWidth:2,
       pointRadius:3,tension:.4,fill:true},
      {label:'Threshold',data:[],borderColor:'#ef4444',borderWidth:2,
       borderDash:[6,4],pointRadius:0,tension:0,fill:false}
    ]},
    options:{responsive:true,maintainAspectRatio:true,animation:{duration:400},
      plugins:{legend:{labels:{color:'#94a3b8',font:{size:12}}}},
      scales:{
        x:{ticks:{color:'#64748b',maxTicksLimit:8,maxRotation:0},
           grid:{color:'rgba(255,255,255,.04)'}},
        y:{beginAtZero:true,ticks:{color:'#64748b',stepSize:1},
           grid:{color:'rgba(255,255,255,.06)'}}
      }}
  });
}

async function updateChart(){
  if(!cameraActive||!crowdChart) return;
  try{
    const d=await fetch('/api/graph-data').then(r=>r.json());
    if(!d.labels.length) return;
    if(graphPh) graphPh.style.display='none';
    crowdChart.data.labels=d.labels;
    crowdChart.data.datasets[0].data=d.counts;
    crowdChart.data.datasets[1].data=d.labels.map(()=>d.threshold);
    crowdChart.data.datasets[0].pointBackgroundColor=d.alerts.map(a=>a?'#ef4444':'#22c55e');
    crowdChart.update();
  }catch{}
}

// ── ALERT SOUND ───────────────────────────────────────────
let isSoundPlaying=false;
function handleAlertSound(alertActive){
  if(!alertSound) return;
  if(alertActive&&!isSoundPlaying){
    alertSound.play().then(()=>{isSoundPlaying=true;}).catch(()=>{
      document.addEventListener('click',()=>{
        if(alertActive){alertSound.play();isSoundPlaying=true;}
      },{once:true});
    });
  } else if(!alertActive&&isSoundPlaying){
    alertSound.pause(); alertSound.currentTime=0; isSoundPlaying=false;
  }
}

let isMuted2=false;
function toggleMute(){
  if(!alertSound) return;
  isMuted2=!isMuted2;
  alertSound.muted=isMuted2;
  document.getElementById('muteBtn').textContent=isMuted2?'🔇 Unmute':'🔊 Mute';
}

// ── POLLING ───────────────────────────────────────────────
function startPolling(){
  pollStatus(); statusInt=setInterval(pollStatus,2000);
  updateChart(); graphInt=setInterval(updateChart,3000);
}

async function pollStatus(){
  try{
    const d=await fetch('/api/status').then(r=>r.json());
    countVal.textContent=d.current_count; threshVal.textContent=d.threshold;
    timeVal.textContent=d.timestamp; flowVal.textContent=d.flow_direction||'—';
    alertBanner.style.display=d.alert_active?'flex':'none';
    handleAlertSound(d.alert_active);

    if(d.camera_active&&!cameraActive){
      cameraActive=true; showFeed(d.camera_source);
    } else if(!d.camera_active&&cameraActive){
      cameraActive=false; hideFeed();
    }
    if(d.camera_active) pollStats();
  }catch{}
}

async function pollStats(){
  try{
    const d=await fetch('/api/session-stats').then(r=>r.json());
    if(d.session_start!=='N/A'&&sessionCard){
      sessionCard.style.display='block';
      document.getElementById('sesStart').textContent=d.session_start;
      document.getElementById('sesDur').textContent=d.session_duration+'s';
      document.getElementById('sesLogs').textContent=d.total_logs;
      document.getElementById('sesAlerts').textContent=d.alert_events;
      document.getElementById('sesPeak').textContent=d.max_crowd;
    }
  }catch{}
}

function showFeed(source){
  feedPh.style.display='none';
  liveStream.style.display='block';
  liveStream.src='/video_feed';
  if(liveTag) liveTag.textContent=source==='esp32'?'Updated every 2s':'Live stream';
}
function hideFeed(){
  liveStream.style.display='none'; liveStream.src='';
  feedPh.style.display='flex';
}

// ── THEME ─────────────────────────────────────────────────
const themeToggle=document.getElementById('themeToggle');
const savedTheme=localStorage.getItem('humix_theme')||'dark';
function applyTheme(t){
  if(t==='light'){document.body.classList.add('light');if(themeToggle)themeToggle.textContent='☀️';}
  else{document.body.classList.remove('light');if(themeToggle)themeToggle.textContent='🌙';}
}
applyTheme(savedTheme);
if(themeToggle) themeToggle.addEventListener('click',()=>{
  const nt=document.body.classList.contains('light')?'dark':'light';
  localStorage.setItem('humix_theme',nt); applyTheme(nt);
});

// ── INIT ──────────────────────────────────────────────────
initChart(); startPolling();
"""

# ── ARDUINO SKETCH ────────────────────────────────────────
files[f"{BASE}/arduino/led_control.ino"] = """\
// HUMIX LED Alert System
// Green  → Pin 8  (Normal)
// Yellow → Pin 9  (Warning: 75% threshold)
// Red    → Pin 10 (Alert: threshold exceeded)

const int GREEN  = 8;
const int YELLOW = 9;
const int RED    = 10;

void setup() {
  Serial.begin(9600);
  pinMode(GREEN,  OUTPUT);
  pinMode(YELLOW, OUTPUT);
  pinMode(RED,    OUTPUT);
  setLED('G'); // Start with green
}

void setLED(char cmd) {
  if (cmd == 'G') {
    digitalWrite(GREEN,  HIGH);
    digitalWrite(YELLOW, LOW);
    digitalWrite(RED,    LOW);
  } else if (cmd == 'Y') {
    digitalWrite(GREEN,  HIGH);
    digitalWrite(YELLOW, HIGH);
    digitalWrite(RED,    LOW);
  } else if (cmd == 'R') {
    digitalWrite(GREEN,  LOW);
    digitalWrite(YELLOW, LOW);
    digitalWrite(RED,    HIGH);
  }
}

void loop() {
  if (Serial.available() > 0) {
    char cmd = Serial.read();
    if (cmd == 'G' || cmd == 'Y' || cmd == 'R') {
      setLED(cmd);
    }
  }
}
"""

# ── WRITE ALL FILES ───────────────────────────────────────
for path, content in files.items():
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)

print("=" * 60)
print("✅  HUMIX System created successfully!")
print(f"📁  Folder: {BASE}/")
print("=" * 60)
print()
print("Steps:")
print(f"  1. Copy yolov8n.pt (or yolo12n.pt)  →  {BASE}/")
print(f"  2. cd {BASE}")
print(f"  3. pip install flask opencv-python ultralytics numpy openpyxl Pillow pyserial")
print(f"  4. python app.py")
print(f"  5. Open http://localhost:5000")
print()
print("User Flow:")
print("  🌐 /           → Landing page (HUMIX)")
print("  📋 /register   → Register your event")
print("  🔑 /login      → Event manager login")
print("  🔐 /admin      → Admin login (admin/admin123)")
print("  📋 /admin/events → Select event to monitor")
print("  📊 Admin dashboard → Camera + image + graph")
print("  🎥 Event dashboard → Live video + graph + alerts")
print()
print("Arduino:")
print("  Upload arduino/led_control.ino to your Arduino")
print("  Connect: Green→Pin8, Yellow→Pin9, Red→Pin10")
print("  Set COM port in Admin dashboard → Connect Arduino")