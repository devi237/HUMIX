

from flask import (Flask, render_template, jsonify, request,
                   session, redirect, url_for, Response, send_file)
import cv2
from ultralytics import YOLO
import time, threading, numpy as np, urllib.request
import sqlite3, os, io, uuid, hashlib
from datetime import datetime
from functools import wraps
mjpeg_reader = None
app = Flask(__name__)
app.secret_key = "humix-secret-2024"

ADMIN_USER     = "admin"
ADMIN_PASS     = "admin123"
DB_PATH        = "database/humix.db"
CAPTURE_INTV   = 2    # seconds between snapshots
IMAGE_INTV     = 180  # 3 minutes image save

# ── GLOBAL STATE ────────────────────────────────────────
model             = None
camera            = None
camera_active     = False
camera_source     = "webcam"
esp32_url         = ""
current_count     = 0
alert_active      = False
flow_direction    = "Analyzing..."
latest_frame      = None
frame_lock        = threading.Lock()
capture_thread    = None
current_event_id  = None
current_threshold = 10
prev_gray         = None
last_image_save   = 0
arduino_serial    = None

# ── Detection runs in its own thread so frame display is never blocked ──
detection_lock    = threading.Lock()
detection_running = False   # True while count_crowd() is executing
pending_raw       = None    # latest raw frame waiting for detection
pending_lock      = threading.Lock()

session_data = {
    "start_time": None, "total_logs": 0,
    "alert_events": 0,  "max_crowd": 0,
    "session_id": None,
}

# ── DATABASE ────────────────────────────────────────────
def init_db():
    os.makedirs("database", exist_ok=True)
    os.makedirs(os.path.join("static","captured_images"), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_name TEXT, date TEXT, start_time TEXT, end_time TEXT,
        venue TEXT, estimated_crowd INTEGER, organizer_name TEXT,
        email TEXT UNIQUE, password TEXT,
        status TEXT DEFAULT 'registered',
        created_at TEXT);
    CREATE TABLE IF NOT EXISTS crowd_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_id INTEGER, session_id TEXT, timestamp TEXT,
        count INTEGER, threshold INTEGER, alert INTEGER, flow_direction TEXT);
    CREATE TABLE IF NOT EXISTS saved_images (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_id INTEGER, session_id TEXT, timestamp TEXT,
        filename TEXT, count INTEGER);
    CREATE TABLE IF NOT EXISTS mon_sessions (
        id TEXT PRIMARY KEY, event_id INTEGER,
        start_time TEXT, end_time TEXT,
        max_crowd INTEGER, total_alerts INTEGER,
        total_logs INTEGER, threshold INTEGER);
    """)
    conn.commit(); conn.close()

def get_db():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c

def hp(p): return hashlib.sha256(p.encode()).hexdigest()

# ── ARDUINO ─────────────────────────────────────────────
def connect_arduino(port):
    global arduino_serial
    try:
        import serial
        arduino_serial = serial.Serial(port, 9600, timeout=1)
        time.sleep(2)
        send_led("G")
        print(f"[HUMIX] Arduino on {port}")
        return True
    except Exception as e:
        print(f"[HUMIX] Arduino error: {e}")
        arduino_serial = None
        return False

def send_led(cmd):
    try:
        if arduino_serial and arduino_serial.is_open:
            arduino_serial.write(f"{cmd}\n".encode())
    except:
        pass

def update_leds(count, threshold):
    if count >= threshold:
        send_led("R")
    elif count >= threshold * 0.75:
        send_led("Y")
    else:
        send_led("G")

# ── MODEL ───────────────────────────────────────────────
def load_model():
    global model
    if model is None:
        print("[HUMIX] Loading model...")
        try:
            model = YOLO("yolo12n.pt")
            print("[HUMIX] YOLOv12 loaded!")
        except:
            model = YOLO("yolov8n.pt")
            print("[HUMIX] YOLOv8 loaded (fallback)")

# ── PREPROCESS ──────────────────────────────────────────
def preprocess(frame):
    h, w = frame.shape[:2]

    # Only upscale if genuinely tiny — never downscale before detection
    if w < 640 or h < 480:
        scale = max(640 / w, 480 / h)
        frame = cv2.resize(frame,
                           (int(w * scale), int(h * scale)),
                           interpolation=cv2.INTER_CUBIC)

    # CLAHE for better contrast in uneven lighting
    lab = cv2.cvtColor(frame, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    l = clahe.apply(l)
    frame = cv2.cvtColor(cv2.merge([l, a, b]), cv2.COLOR_LAB2BGR)

    # Mild sharpening
    blurred = cv2.GaussianBlur(frame, (0, 0), 2)
    return cv2.addWeighted(frame, 1.3, blurred, -0.3, 0)

# ── DENSITY ESTIMATION (fallback for crowded scenes) ─────
def estimate_density(frame):
    """
    Uses edge density + skin-tone blob analysis to estimate
    number of people when YOLO misses partially occluded bodies.
    Returns an estimated additional crowd count.
    """
    try:
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        # Edge density analysis — more edges = more bodies
        edges = cv2.Canny(gray, 50, 150)
        edge_ratio = np.count_nonzero(edges) / edges.size

        # Skin-tone blob detection in HSV
        hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
        lower_skin = np.array([0, 20, 70], dtype=np.uint8)
        upper_skin = np.array([20, 255, 255], dtype=np.uint8)
        skin_mask = cv2.inRange(hsv, lower_skin, upper_skin)
        # Also capture darker skin tones
        lower_skin2 = np.array([170, 20, 70], dtype=np.uint8)
        upper_skin2 = np.array([180, 255, 255], dtype=np.uint8)
        skin_mask2 = cv2.inRange(hsv, lower_skin2, upper_skin2)
        skin_mask = cv2.bitwise_or(skin_mask, skin_mask2)
        skin_mask = cv2.morphologyEx(skin_mask, cv2.MORPH_OPEN,
                                     np.ones((5,5), np.uint8))
        skin_mask = cv2.morphologyEx(skin_mask, cv2.MORPH_DILATE,
                                     np.ones((10,10), np.uint8))
        contours, _ = cv2.findContours(skin_mask, cv2.RETR_EXTERNAL,
                                        cv2.CHAIN_APPROX_SIMPLE)
        # Each sizable skin blob likely represents a face/neck area
        skin_blobs = sum(1 for c in contours
                         if 200 < cv2.contourArea(c) < 20000)
        return skin_blobs, edge_ratio
    except:
        return 0, 0.0

# ── SMART CROWD COUNT ────────────────────────────────────
def count_crowd(frame):
    h, w = frame.shape[:2]
    display = frame.copy()
    all_boxes = []

    try:
        # Pass 1: Full frame — catches full/partial bodies
        for r in model(frame, stream=True, conf=0.35, iou=0.45,
                       classes=[0], verbose=False):
            for box in r.boxes:
                x1, y1, x2, y2 = map(int, box.xyxy[0])
                conf = float(box.conf[0])
                all_boxes.append((x1, y1, x2, y2, conf, "full"))

        # Pass 2: Top 65% crop — catches heads/torsos in crowds
        crop_h = int(h * 0.65)
        top_crop = frame[0:crop_h, 0:w]
        for r in model(top_crop, stream=True, conf=0.30, iou=0.35,
                       classes=[0], verbose=False):
            for box in r.boxes:
                x1, y1, x2, y2 = map(int, box.xyxy[0])
                conf = float(box.conf[0])
                box_area = (x2 - x1) * (y2 - y1)
                if box_area == 0:
                    continue
                is_dup = any(
                    max(0, min(x2, bx2) - max(x1, bx1)) *
                    max(0, min(y2, by2) - max(y1, by1)) > 0.4 * box_area
                    for (bx1, by1, bx2, by2, _, _) in all_boxes
                )
                if not is_dup:
                    all_boxes.append((x1, y1, x2, y2, conf, "upper"))

    except Exception as e:
        print(f"[HUMIX] Detection error: {e}")

    yolo_count = len(all_boxes)

    # Density cross-check
    skin_blobs, edge_ratio = estimate_density(frame)
    density_count = skin_blobs
    if density_count > yolo_count * 1.5 and edge_ratio > 0.08:
        final_count = int(yolo_count * 0.6 + density_count * 0.4)
    else:
        final_count = yolo_count

    colors = {"full": (0, 220, 100), "upper": (0, 180, 255)}
    for (x1, y1, x2, y2, conf, src) in all_boxes:
        color = colors.get(src, (0, 220, 100))
        cv2.rectangle(display, (x1, y1), (x2, y2), color, 2)
        cv2.putText(display, f"{conf:.2f}",
                    (x1, max(y1 - 6, 10)),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.38, color, 1)

    return final_count, display, all_boxes
# ── OPTICAL FLOW ────────────────────────────────────────
def get_flow(frame):
    global prev_gray
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    d = "Still"
    if prev_gray is not None and prev_gray.shape == gray.shape:
        try:
            flow = cv2.calcOpticalFlowFarneback(
                prev_gray,gray,None,0.5,3,15,3,5,1.2,0)
            fx,fy = flow[...,0].mean(), flow[...,1].mean()
            mag = (fx**2+fy**2)**0.5
            if mag > 0.3:
                if abs(fx)>abs(fy): d="Moving RIGHT →" if fx>0 else "Moving LEFT ←"
                else:               d="Moving DOWN ↓"  if fy>0 else "Moving UP ↑"
        except: d="Analyzing..."
    prev_gray = gray.copy()
    return d

# ── MJPEG STREAM (event page) ───────────────────────────
def generate_stream():
    while camera_active:
        if camera_source == "webcam" and camera and camera.isOpened():
            ok, frame = camera.read()
            if ok and frame is not None:
                cv2.putText(frame, f"People: {current_count}", (10,30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,220,100), 2)
                cv2.putText(frame, f"Threshold: {current_threshold}", (10,58),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 1)
                cv2.putText(frame, flow_direction, (10,86),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.55, (200,200,0), 1)
                if alert_active:
                    cv2.rectangle(frame,(0,0),(frame.shape[1]-1,frame.shape[0]-1),(0,0,255),4)
                    cv2.putText(frame,"ALERT: CROWD LIMIT EXCEEDED",
                        (10,frame.shape[0]-15),cv2.FONT_HERSHEY_SIMPLEX,0.65,(0,0,255),2)
                ret, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY,70])
                if ret:
                    yield (b"--frame\r\nContent-Type: image/jpeg\r\n\r\n"
                           + buf.tobytes() + b"\r\n")

        elif camera_source == "esp32" and esp32_url:
            # ✅ Use ESP32 /stream endpoint directly for live video
            try:
                stream_url = esp32_url.replace("/capture", "/stream")
                resp = urllib.request.urlopen(stream_url, timeout=10)
                boundary = b"--frame"
                buffer = b""
                while camera_active:
                    chunk = resp.read(1024)
                    if not chunk: break
                    buffer += chunk
                    start = buffer.find(b"\xff\xd8")  # JPEG start
                    end   = buffer.find(b"\xff\xd9")  # JPEG end
                    if start != -1 and end != -1:
                        jpg = buffer[start:end+2]
                        buffer = buffer[end+2:]
                        arr  = np.frombuffer(jpg, dtype=np.uint8)
                        frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)
                        if frame is not None:
                            # Add overlay
                            cv2.putText(frame, f"People: {current_count}", (10,30),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,220,100), 2)
                            cv2.putText(frame, f"Threshold: {current_threshold}", (10,58),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 1)
                            if alert_active:
                                cv2.rectangle(frame,(0,0),
                                    (frame.shape[1]-1,frame.shape[0]-1),(0,0,255),4)
                            ret, buf = cv2.imencode(".jpg", frame,
                                [cv2.IMWRITE_JPEG_QUALITY,70])
                            if ret:
                                yield (b"--frame\r\nContent-Type: image/jpeg\r\n\r\n"
                                       + buf.tobytes() + b"\r\n")
            except Exception as e:
                print(f"[HUMIX] Stream error: {e}")
                time.sleep(2)

        time.sleep(0.033)

# ── DETECTION WORKER (runs in its own thread) ───────────
def detection_worker():
    global current_count, alert_active, latest_frame, flow_direction
    global session_data, last_image_save, current_threshold, pending_raw, detection_running

    print("[DETECT] Thread started")
    last_image_save = time.time()

    while camera_active:
        print(f"[DETECT] alive | camera_active={camera_active}")

        with pending_lock:
            raw = pending_raw
            pending_raw = None

        if raw is None:
            time.sleep(0.05)
            continue

        detection_running = True
        try:
            try:
                frame = preprocess(raw)
            except Exception as e:
                print(f"[DETECT] Preprocess error: {e}")
                frame = raw.copy()

            try:
                flow_direction = get_flow(frame)
            except:
                flow_direction = "Analyzing..."

            count = 0
            display = frame.copy()
            try:
                count, display, _ = count_crowd(frame)
            except Exception as e:
                print(f"[HUMIX] Detect: {e}")

            prev = alert_active
            current_count = count
            alert_active = count > current_threshold
            if alert_active and not prev:
                session_data["alert_events"] += 1
            session_data["max_crowd"] = max(session_data["max_crowd"], count)
            session_data["total_logs"] += 1

            update_leds(count, current_threshold)

            sid = session_data["session_id"]
            if current_event_id and sid:
                try:
                    conn = get_db()
                    conn.execute("INSERT INTO crowd_logs VALUES (NULL,?,?,?,?,?,?,?)",
                                 (current_event_id, sid,
                                  datetime.now().strftime("%H:%M:%S"),
                                  count, current_threshold,
                                  1 if alert_active else 0, flow_direction))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    print(f"[DETECT] DB log error: {e}")

            now = time.time()
            if now - last_image_save >= IMAGE_INTV and current_event_id:
                try:
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    fn = f"ev{current_event_id}_{ts}.jpg"
                    fp = os.path.join("static", "captured_images", fn)
                    cv2.imwrite(fp, display)
                    conn = get_db()
                    conn.execute("INSERT INTO saved_images VALUES (NULL,?,?,?,?,?)",
                                 (current_event_id, sid,
                                  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                  fn, count))
                    conn.commit()
                    conn.close()
                    last_image_save = now
                    print(f"[HUMIX] Image saved: {fn}")
                except Exception as e:
                    print(f"[HUMIX] Image save error: {e}")

            # ── HUD overlay ──
            h_f = display.shape[0]
            cv2.rectangle(display, (0, 0), (310, 120), (0, 0, 0), -1)
            cv2.putText(display, f"People : {count}",
                        (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255, 255, 255), 2)
            cv2.putText(display, f"Threshold: {current_threshold}",
                        (10, 58), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (180, 180, 180), 1)
            cv2.putText(display, datetime.now().strftime("%H:%M:%S"),
                        (10, 84), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (140, 140, 140), 1)
            cv2.putText(display, flow_direction,
                        (10, 110), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (200, 200, 0), 1)
            if alert_active:
                cv2.rectangle(display, (0, 0),
                              (display.shape[1] - 1, h_f - 1), (0, 0, 230), 5)
                cv2.putText(display, "ALERT: CROWD LIMIT EXCEEDED",
                            (10, h_f - 15), cv2.FONT_HERSHEY_SIMPLEX,
                            0.65, (0, 0, 255), 2)

            with frame_lock:
                latest_frame = display.copy()

            print(f"[HUMIX] Count:{count} | Flow:{flow_direction}")

        except Exception as e:
            print(f"[DETECT] Worker error: {e}")
            import traceback
            traceback.print_exc()
        finally:
            detection_running = False

    print("[DETECT] Thread exiting")

# ── CAPTURE WORKER (grabs frames, pushes raw instantly) ──
def capture_worker():
    global latest_frame, pending_raw

    print(f"[CAPTURE] Thread started | source={camera_source} | url={esp32_url} | mjpeg_reader={mjpeg_reader}")
    last_capture = 0
    time.sleep(2.0)  # grace period — let ESP32 stream settle
    print("[CAPTURE] Grace period done, entering loop")

    while camera_active:
        print(f"[CAPTURE] alive | camera_active={camera_active}")
        now = time.time()
        if last_capture != 0 and (now - last_capture) < CAPTURE_INTV:
            time.sleep(0.03)
            continue

        raw = None
        try:
            if camera_source == "webcam":
                if camera and camera.isOpened():
                    ok, raw = False, None
                    for _ in range(3):
                        ok, raw = camera.read()
                    if not ok or raw is None:
                        time.sleep(0.5)
                        continue
                    if raw.mean() < 10:
                        time.sleep(0.3)
                        continue

            elif camera_source == "esp32":
                if mjpeg_reader is None:
                    print("[CAPTURE] mjpeg_reader is None — waiting...")
                    time.sleep(1)
                    continue
                raw = mjpeg_reader.get_frame()
                if raw is None:
                    time.sleep(0.1)
                    continue
                h_raw, w_raw = raw.shape[:2]
                print(f"[HUMIX] ESP32 frame: {w_raw}x{h_raw} | mean={raw.mean():.1f}")

            if raw is None:
                continue

            # ── Preview: use native res, only downscale if > 1280px ──
            try:
                h_r, w_r = raw.shape[:2]
                if w_r > 1280:
                    scale = 1280 / w_r
                    preview = cv2.resize(raw,
                                         (1280, int(h_r * scale)),
                                         interpolation=cv2.INTER_AREA)
                else:
                    preview = raw.copy()

                if latest_frame is None:
                    cv2.putText(preview, "Initializing detection...",
                                (10, 30), cv2.FONT_HERSHEY_SIMPLEX,
                                0.7, (0, 220, 100), 2)
                with frame_lock:
                    latest_frame = preview.copy()
            except Exception as e:
                print(f"[CAPTURE] Preview error: {e}")

            # ── Queue full-res frame for detection ──
            with pending_lock:
                pending_raw = raw

            last_capture = now

        except Exception as e:
            print(f"[CAPTURE] Error: {e}")
            import traceback
            traceback.print_exc()
            time.sleep(0.5)

    print("[CAPTURE] Thread exiting")

# ── CAMERA MANAGEMENT ───────────────────────────────────
def start_camera(source, url="", event_id=None, threshold=10):
    global camera, camera_active, camera_source, esp32_url
    global capture_thread, session_data, current_event_id, current_threshold
    global prev_gray, latest_frame, pending_raw, mjpeg_reader

    print(f"[HUMIX] start_camera called | source={source} | url={url} | event_id={event_id}")

    # ── Force-reset any stale state from previous session ──
    if camera_active:
        dead = False
        if camera_source == "esp32":
            dead = (mjpeg_reader is None or mjpeg_reader.get_frame() is None)
        elif camera_source == "webcam":
            dead = (camera is None or not camera.isOpened())

        if dead:
            print("[HUMIX] Stale session detected — force resetting...")
            camera_active = False
            if mjpeg_reader:
                mjpeg_reader.stop()
                mjpeg_reader = None
            if camera:
                camera.release()
                camera = None
            time.sleep(0.5)  # let old threads exit
        else:
            print("[HUMIX] Already running")
            return False

    # ── Reset frame state ──
    with frame_lock:
        latest_frame = None
    with pending_lock:
        pending_raw = None
    prev_gray = None

    # ── Set globals ──
    camera_source = source
    esp32_url = url
    current_event_id = event_id
    current_threshold = threshold

    # ── Source-specific setup ──
    if source == "webcam":
        camera = cv2.VideoCapture(0)
        camera.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        camera.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        if not camera.isOpened():
            print("[HUMIX] Webcam failed to open")
            return False
        for _ in range(20):
            camera.read()  # flush warm-up frames
        print("[HUMIX] Webcam ready ✓")

    elif source == "esp32":
        if not url:
            print("[HUMIX] ESP32 source selected but no URL provided")
            return False
        stream_url = url if "/stream" in url else url + "/stream"
        print(f"[HUMIX] Starting MJPEG reader for {stream_url}")
        mjpeg_reader = MJPEGStreamReader(stream_url)
        mjpeg_reader.start()
        # Wait up to 5s for first valid frame
        print("[HUMIX] Waiting for ESP32 first frame...")
        for _ in range(50):
            if mjpeg_reader.get_frame() is not None:
                print("[HUMIX] ESP32 stream ready ✓")
                break
            time.sleep(0.1)
        else:
            print("[HUMIX] Warning: no frame yet, continuing anyway")

    else:
        print(f"[HUMIX] Unknown source: {source}")
        return False

    # ── Load model ──
    load_model()

    # ── Session setup ──
    sid = str(uuid.uuid4())[:8]
    camera_active = True  # set AFTER all setup

    session_data.update({
        "start_time": datetime.now(),
        "total_logs": 0,
        "alert_events": 0,
        "max_crowd": 0,
        "session_id": sid
    })

    if event_id:
        try:
            conn = get_db()
            conn.execute("INSERT INTO mon_sessions VALUES (?,?,?,NULL,0,0,0,?)",
                         (sid, event_id,
                          datetime.now().strftime("%Y-%m-%d %H:%M:%S"), threshold))
            conn.execute("UPDATE events SET status='active' WHERE id=?", (event_id,))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"[HUMIX] DB error on session insert: {e}")

    # ── Launch worker threads ──
    print(f"[HUMIX] Launching capture_thread | source={source}")
    capture_thread = threading.Thread(target=capture_worker, daemon=True)
    capture_thread.start()
    print(f"[HUMIX] capture_thread alive: {capture_thread.is_alive()}")

    det_thread = threading.Thread(target=detection_worker, daemon=True)
    det_thread.start()
    print(f"[HUMIX] detection_thread alive: {det_thread.is_alive()}")

    print(f"[HUMIX] Session {sid} started | source={source} | threshold={threshold}")
    return True


# ── STOP CAMERA ────────────────────────────────────────────────────────────
def stop_camera():
    global camera, camera_active, mjpeg_reader

    sid = session_data.get("session_id")
    if sid and current_event_id:
        try:
            conn = get_db()
            conn.execute(
                "UPDATE mon_sessions SET end_time=?,max_crowd=?,total_alerts=?,total_logs=? WHERE id=?",
                (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                 session_data["max_crowd"], session_data["alert_events"],
                 session_data["total_logs"], sid))
            conn.execute("UPDATE events SET status='completed' WHERE id=?",
                         (current_event_id,))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"[HUMIX] DB error on stop: {e}")

    send_led("G")
    camera_active = False

    if mjpeg_reader:
        mjpeg_reader.stop()
        mjpeg_reader = None
        print("[HUMIX] MJPEG reader stopped")

    if camera:
        camera.release()
        camera = None
        print("[HUMIX] Webcam released")

    print("[HUMIX] Camera stopped")
    
    #------------------------------------------------- 
class MJPEGStreamReader:
    """
    Single persistent connection to ESP32 MJPEG stream.
    Reconnects automatically on drop. Thread-safe frame access.
    """
    def __init__(self, url):
        self.url = url
        self.latest_frame = None
        self.running = False
        self._lock = threading.Lock()
        self._thread = None

    def start(self):
        self.running = True
        self._thread = threading.Thread(target=self._reader_loop, daemon=True)
        self._thread.start()

    def stop(self):
        self.running = False

    def get_frame(self):
        with self._lock:
            return self.latest_frame.copy() if self.latest_frame is not None else None

    def _reader_loop(self):
        global camera_active
        while self.running:
            try:
                print(f"[MJPEG] Connecting to {self.url}")
                resp = urllib.request.urlopen(self.url, timeout=15)
                buffer = b""
                consecutive_black = 0

                while self.running:
                    chunk = resp.read(8192)
                    if not chunk:
                        break
                    buffer += chunk

                    # Prevent buffer growing unbounded
                    if len(buffer) > 500_000:
                        buffer = buffer[-200_000:]

                    start = buffer.find(b'\xff\xd8')  # JPEG start
                    end   = buffer.find(b'\xff\xd9')  # JPEG end

                    if start != -1 and end != -1 and end > start:
                        jpg = buffer[start:end + 2]
                        buffer = buffer[end + 2:]

                        arr = np.frombuffer(jpg, dtype=np.uint8)
                        frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)

                        if frame is None:
                            continue

                        if frame.mean() < 10:
                            consecutive_black += 1
                            if consecutive_black > 10:
                                print("[MJPEG] Too many black frames, reconnecting")
                                break
                            continue

                        consecutive_black = 0
                        with self._lock:
                            self.latest_frame = frame

            except Exception as e:
                print(f"[MJPEG] Stream error: {e}")

            if self.running:
                print("[MJPEG] Reconnecting in 3s...")
                time.sleep(3)

        # Stream fully stopped
        camera_active = False
        print("[MJPEG] Reader stopped")

# ── AUTH DECORATORS ─────────────────────────────────────
def admin_req(f):
    @wraps(f)
    def d(*a,**k):
        if not session.get("admin"): return redirect(url_for("admin_login_page"))
        return f(*a,**k)
    return d

def event_req(f):
    @wraps(f)
    def d(*a,**k):
        if not session.get("event_id"): return redirect(url_for("event_login_page"))
        return f(*a,**k)
    return d

def api_auth(f):
    @wraps(f)
    def d(*a,**k):
        if not session.get("admin") and not session.get("event_id"):
            return jsonify({"error":"Unauthorized"}),401
        return f(*a,**k)
    return d

def api_admin(f):
    @wraps(f)
    def d(*a,**k):
        if not session.get("admin"): return jsonify({"error":"Admin only"}),403
        return f(*a,**k)
    return d
def _reader_loop(self):
    global camera_active  # ← add this
    while self.running:
        try:
            # ... existing loop code ...
            pass
        except Exception as e:
            print(f"[MJPEG] Stream error: {e}")

        if self.running:
            print("[MJPEG] Reconnecting in 3s...")
            time.sleep(3)
    
    # Stream fully stopped — clear flag so UI can restart
    camera_active = False  # ← add this at the end
# ── ROUTES ──────────────────────────────────────────────
@app.route("/")
def landing():
    return render_template("landing.html")

@app.route("/register", methods=["GET","POST"])
def register():
    if request.method=="POST":
        d=request.form
        try:
            conn=get_db()
            conn.execute("""INSERT INTO events
                (event_name,date,start_time,end_time,venue,estimated_crowd,
                 organizer_name,email,password,status,created_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                (d.get("event_name"),d.get("date"),d.get("start_time"),
                 d.get("end_time"),d.get("venue"),int(d.get("estimated_crowd",0)),
                 d.get("organizer_name"),d.get("email"),hp(d.get("password","")),
                 "registered",datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit(); conn.close()
            return redirect(url_for("event_login_page")+"?registered=1")
        except sqlite3.IntegrityError:
            return render_template("register.html",error="Email already registered")
        except Exception as e:
            return render_template("register.html",error=str(e))
    return render_template("register.html")

@app.route("/login", methods=["GET","POST"])
def event_login_page():
    registered=request.args.get("registered")
    if request.method=="POST":
        conn=get_db()
        ev=conn.execute("SELECT * FROM events WHERE email=? AND password=?",
            (request.form.get("email"),hp(request.form.get("password","")))).fetchone()
        conn.close()
        if ev:
            session["event_id"]=ev["id"]
            session["event_name"]=ev["event_name"]
            return redirect(url_for("event_dashboard"))
        return render_template("event_login.html",error="Invalid credentials")
    return render_template("event_login.html",registered=registered)

@app.route("/admin")
def admin_login_page():
    if session.get("admin"): return redirect(url_for("admin_events"))
    return render_template("admin_login.html")

@app.route("/admin/login", methods=["POST"])
def admin_login():
    if (request.form.get("username")==ADMIN_USER and
        request.form.get("password")==ADMIN_PASS):
        session["admin"]=True
        return redirect(url_for("admin_events"))
    return render_template("admin_login.html",error="Invalid admin credentials")

@app.route("/admin/events")
@admin_req
def admin_events():
    conn=get_db()
    events=conn.execute("SELECT * FROM events ORDER BY created_at DESC").fetchall()
    conn.close()
    return render_template("admin_events.html",events=events)

@app.route("/admin/select/<int:eid>")
@admin_req
def admin_select(eid):
    conn=get_db()
    ev=conn.execute("SELECT * FROM events WHERE id=?",(eid,)).fetchone()
    conn.close()
    if ev:
        session["sel_eid"]=eid
        session["sel_ename"]=ev["event_name"]
        session["sel_thresh"]=ev["estimated_crowd"]
        return redirect(url_for("admin_dashboard"))
    return redirect(url_for("admin_events"))

@app.route("/admin/dashboard")
@admin_req
def admin_dashboard():
    if not session.get("sel_eid"): return redirect(url_for("admin_events"))
    conn=get_db()
    ev=conn.execute("SELECT * FROM events WHERE id=?",(session["sel_eid"],)).fetchone()
    conn.close()
    return render_template("admin_dashboard.html",event=ev,camera_active=camera_active)

@app.route("/event/dashboard")
@event_req
def event_dashboard():
    conn=get_db()
    ev=conn.execute("SELECT * FROM events WHERE id=?",(session["event_id"],)).fetchone()
    conn.close()
    return render_template("event_dashboard.html",event=ev)

@app.route("/logout")
def logout():
    session.clear(); return redirect(url_for("landing"))

@app.route("/admin/logout")
def admin_logout():
    session.clear(); stop_camera(); return redirect(url_for("landing"))

# ── API ─────────────────────────────────────────────────
@app.route("/latest_image")
@api_auth
def latest_image():
    if not camera_active or latest_frame is None: return "No image",404
    for _ in range(50):
        with frame_lock: f=latest_frame
        if f is not None: break
        time.sleep(0.1)
    else: return "Not ready",404
    with frame_lock:
        ok,buf=cv2.imencode(".jpg",latest_frame,[cv2.IMWRITE_JPEG_QUALITY,85])
    if not ok: return "Error",500
    return buf.tobytes(),200,{"Content-Type":"image/jpeg","Cache-Control":"no-store"}

@app.route("/video_feed")
def video_feed():
    def generate():
        while True:
            frame = None

            # Priority 1: raw frames from ESP32 via shared reader (smooth, fast)
            if camera_source == "esp32" and mjpeg_reader is not None:
                frame = mjpeg_reader.get_frame()

            # Priority 2: latest annotated frame (webcam or fallback)
            if frame is None:
                with frame_lock:
                    if latest_frame is not None:
                        frame = latest_frame.copy()

            if frame is None:
                time.sleep(0.05)
                continue

            _, jpg = cv2.imencode(".jpg", frame,
                                  [cv2.IMWRITE_JPEG_QUALITY, 75])
            yield (b"--frame\r\nContent-Type: image/jpeg\r\n\r\n"
                   + jpg.tobytes() + b"\r\n")
            time.sleep(0.04)  # ~25fps

    return Response(generate(),
                    mimetype="multipart/x-mixed-replace; boundary=frame")

@app.route("/api/camera/start",methods=["POST"])
@api_admin
def api_start():
    d=request.json or {}
    if camera_active: return jsonify({"success":False,"message":"Already running"})
    ok=start_camera(d.get("source","webcam"),d.get("url",""),
        event_id=session.get("sel_eid"),
        threshold=session.get("sel_thresh",10))
    return jsonify({"success":ok,"message":"Started" if ok else "Failed"})

@app.route("/api/camera/stop",methods=["POST"])
@api_admin
def api_stop():
    stop_camera()
    return jsonify({"success":True})

@app.route("/api/status")
@api_auth
def api_status():
    return jsonify({
        "camera_active":camera_active,"camera_source":camera_source,
        "current_count":current_count,"alert_active":alert_active,
        "threshold":current_threshold,"flow_direction":flow_direction,
        "timestamp":datetime.now().strftime("%H:%M:%S"),
    })

@app.route("/api/session-stats")
@api_auth
def api_stats():
    if not session_data["start_time"]:
        return jsonify({"session_start":"N/A","session_duration":0,
                        "total_logs":0,"alert_events":0,"max_crowd":0})
    dur=int((datetime.now()-session_data["start_time"]).total_seconds())
    return jsonify({
        "session_start":session_data["start_time"].strftime("%H:%M:%S"),
        "session_duration":dur,"total_logs":session_data["total_logs"],
        "alert_events":session_data["alert_events"],"max_crowd":session_data["max_crowd"],
    })

@app.route("/api/graph-data")
@api_auth
def api_graph():
    eid=session.get("sel_eid") or session.get("event_id")
    sid=session_data.get("session_id")
    try:
        conn=get_db()
        rows=conn.execute(
            "SELECT timestamp,count,alert FROM crowd_logs WHERE event_id=? AND session_id=? ORDER BY id DESC LIMIT 30",
            (eid,sid)).fetchall()
        conn.close()
        rows=list(reversed(rows))
        return jsonify({"labels":[r["timestamp"] for r in rows],
                        "counts":[r["count"] for r in rows],
                        "alerts":[r["alert"] for r in rows],
                        "threshold":current_threshold})
    except:
        return jsonify({"labels":[],"counts":[],"alerts":[],"threshold":current_threshold})

@app.route("/api/settings/threshold",methods=["POST"])
@api_admin
def api_thresh():
    global current_threshold
    v=(request.json or {}).get("threshold")
    if v and isinstance(v,int) and v>0:
        current_threshold=v
        session["sel_thresh"]=v
        return jsonify({"success":True,"message":f"Threshold set to {v}"})
    return jsonify({"success":False,"message":"Invalid"})

@app.route("/api/settings/confidence",methods=["POST"])
@api_admin
def api_conf():
    return jsonify({"success":True})

@app.route("/api/arduino/connect",methods=["POST"])
@api_admin
def api_arduino():
    port=(request.json or {}).get("port","COM3")
    ok=connect_arduino(port)
    return jsonify({"success":ok,"message":f"Connected on {port}" if ok else "Failed"})

@app.route("/api/report")
@api_auth
def api_report():
    try:
        import openpyxl
        from openpyxl.styles import Font,PatternFill,Alignment
        from openpyxl.chart import LineChart,Reference
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImg

        eid=session.get("sel_eid") or session.get("event_id")
        sid=session_data.get("session_id")
        conn=get_db()
        ev  =conn.execute("SELECT * FROM events WHERE id=?",(eid,)).fetchone()
        logs=conn.execute(
            "SELECT timestamp,count,threshold,alert,flow_direction FROM crowd_logs WHERE event_id=? AND session_id=? ORDER BY id",
            (eid,sid)).fetchall()
        imgs=conn.execute(
            "SELECT timestamp,filename,count FROM saved_images WHERE event_id=? AND session_id=? ORDER BY id LIMIT 3",
            (eid,sid)).fetchall()
        conn.close()

        wb=openpyxl.Workbook()

        # Sheet 1 – Summary
        ws=wb.active; ws.title="Event Summary"
        ws.merge_cells("A1:E1")
        ws["A1"]="HUMIX — Smart Human Density & Movement Monitoring System"
        ws["A1"].font=Font(size=14,bold=True,color="FFFFFF")
        ws["A1"].fill=PatternFill("solid",fgColor="0c0e14")
        ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[1].height=40

        ws.merge_cells("A2:E2")
        ws["A2"]=f"Event Report: {ev['event_name'] if ev else 'N/A'}"
        ws["A2"].font=Font(size=12,bold=True,color="22c55e")
        ws["A2"].fill=PatternFill("solid",fgColor="13161f")
        ws["A2"].alignment=Alignment(horizontal="center")
        ws.row_dimensions[2].height=28

        info=[
            ("",""),("EVENT DETAILS",""),
            ("Event Name",   ev["event_name"]      if ev else ""),
            ("Date",         ev["date"]             if ev else ""),
            ("Start Time",   ev["start_time"]       if ev else ""),
            ("End Time",     ev["end_time"]         if ev else ""),
            ("Venue",        ev["venue"]            if ev else ""),
            ("Organizer",    ev["organizer_name"]   if ev else ""),
            ("Est. Crowd",   ev["estimated_crowd"]  if ev else ""),
            ("",""),("MONITORING SUMMARY",""),
            ("Session Start", session_data["start_time"].strftime("%Y-%m-%d %H:%M:%S") if session_data["start_time"] else "N/A"),
            ("Peak Count",    session_data["max_crowd"]),
            ("Alert Events",  session_data["alert_events"]),
            ("Total Logs",    session_data["total_logs"]),
            ("Threshold",     current_threshold),
            ("Alert Status",  "ALERTS TRIGGERED" if session_data["alert_events"]>0 else "NO ALERTS"),
        ]
        for i,(lbl,val) in enumerate(info,3):
            if lbl in ("EVENT DETAILS","MONITORING SUMMARY"):
                ws.merge_cells(f"A{i}:E{i}")
                ws[f"A{i}"]=lbl
                ws[f"A{i}"].font=Font(bold=True,color="FFFFFF",size=10)
                ws[f"A{i}"].fill=PatternFill("solid",fgColor="22c55e")
            elif lbl:
                ws[f"A{i}"]=lbl; ws[f"B{i}"]=str(val)
                ws[f"A{i}"].font=Font(bold=True)
        ws.column_dimensions["A"].width=22
        ws.column_dimensions["B"].width=35

        # Sheet 2 – Data Log
        ws2=wb.create_sheet("Crowd Data Log")
        hdrs=["Time","Count","Threshold","Alert","Flow Direction"]
        gf=PatternFill("solid",fgColor="22c55e")
        rf=PatternFill("solid",fgColor="FEE2E2")
        for col,h in enumerate(hdrs,1):
            c=ws2.cell(1,col,h)
            c.font=Font(bold=True,color="FFFFFF"); c.fill=gf
            c.alignment=Alignment(horizontal="center")
            ws2.column_dimensions[get_column_letter(col)].width=18
        for rn,row in enumerate(logs,2):
            ws2.cell(rn,1,row["timestamp"]); ws2.cell(rn,2,row["count"])
            ws2.cell(rn,3,row["threshold"])
            ws2.cell(rn,4,"YES" if row["alert"] else "No")
            ws2.cell(rn,5,row["flow_direction"])
            if row["alert"]:
                for col in range(1,6): ws2.cell(rn,col).fill=rf

        # Sheet 3 – Graph
        ws3=wb.create_sheet("Graph")
        ws3["A1"],ws3["B1"],ws3["C1"]="Time","Count","Threshold"
        for i,row in enumerate(logs,2):
            ws3.cell(i,1,row["timestamp"])
            ws3.cell(i,2,row["count"])
            ws3.cell(i,3,row["threshold"])
        if len(logs)>1:
            ch=LineChart()
            ch.title="Crowd Count Over Time"
            ch.style=10; ch.width=28; ch.height=16
            ch.y_axis.title="People"; ch.x_axis.title="Time"
            dr=Reference(ws3,min_col=2,max_col=3,min_row=1,max_row=len(logs)+1)
            lr=Reference(ws3,min_col=1,min_row=2,max_row=len(logs)+1)
            ch.add_data(dr,titles_from_data=True); ch.set_categories(lr)
            ws3.add_chart(ch,"E2")

        # Sheet 4 – Captured Images (up to 3)
        if imgs:
            ws4=wb.create_sheet("Captured Images")
            ws4["A1"]="Captured Monitoring Images (every 3 minutes)"
            ws4["A1"].font=Font(size=13,bold=True)
            ws4.row_dimensions[1].height=28
            rp=3
            for img in imgs:
                fp=os.path.join("static","captured_images",img["filename"])
                ws4.cell(rp,1,f"Time: {img['timestamp']}  |  Count: {img['count']}")
                ws4.cell(rp,1).font=Font(bold=True,color="22c55e")
                rp+=1
                if os.path.exists(fp):
                    try:
                        xi=XLImg(fp); xi.width=420; xi.height=290
                        ws4.add_image(xi,f"A{rp}"); rp+=17
                    except:
                        ws4.cell(rp,1,f"[{img['filename']}]"); rp+=2
                else:
                    rp+=1

        out=io.BytesIO(); wb.save(out); out.seek(0)
        fn=f"HUMIX_{ev['event_name'].replace(' ','_') if ev else 'Report'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,download_name=fn)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error":str(e)}),500
def connect_arduino(port):
    global arduino_serial
    try:
        import serial
        print(f"[HUMIX] Trying to connect Arduino on {port}...")
        arduino_serial = serial.Serial(port, 9600, timeout=1)
        time.sleep(2)
        send_led("G")
        print(f"[HUMIX] Arduino connected successfully on {port}")
        return True
    except serial.SerialException as e:
        print(f"[HUMIX] SerialException: {e}")
        arduino_serial = None
        return False
    except Exception as e:
        print(f"[HUMIX] Error: {e}")
        arduino_serial = None
        return False
if __name__=="__main__":
    init_db()
    app.run(debug=True,host="0.0.0.0",port=5000)
